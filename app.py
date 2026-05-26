"""
PULSO DE CALIDAD SPC
Aplicacion web para Control Estadistico de Procesos (CEP/SPC).
Desarrollador: Jerson Andres Lopez Wilches
Contacto: jerssonpriv@gmail.com

Uso recomendado:
  NUBE STREAMLIT BUSCAR CON NOMBRE PULSO DE CALIDAD SPC+

Estructura logica del codigo:
    1. Dependencias y utilidades generales.
    2. Banco central de parametros del proceso.
    3. Calculos estadisticos CEP.
    4. Graficos y componentes visuales.
    5. Pantallas o secciones de la aplicacion.
    6. Pruebas internas y arranque de Streamlit.
"""

from __future__ import annotations

import sys
import time
import subprocess
import importlib.util
from io import StringIO, BytesIO
from typing import Optional
import tempfile
import os

PAQUETES_APP = [
    "streamlit",
    "pandas",
    "numpy",
    "plotly",
    "scipy",
    "statsmodels",
    "openpyxl",
]

PAQUETES_BASE = ["pandas", "numpy", "scipy", "statsmodels"]


def existe_paquete(nombre: str) -> bool:
    return importlib.util.find_spec(nombre) is not None


def paquetes_faltantes(paquetes: list[str]) -> list[str]:
    return [p for p in paquetes if not existe_paquete(p)]


def instalar_paquetes(paquetes: list[str]) -> bool:
    if not paquetes:
        print("Todas las librerías ya están instaladas.")
        return True
    comando = [sys.executable, "-m", "pip", "install", *paquetes]
    print("Instalando librerías:", ", ".join(paquetes))
    try:
        subprocess.check_call(comando)
        print("Instalación completada.")
        print("Ejecuta: py -m streamlit run PulsoCalidadSPC_FINAL_DEPURADA.py")
        return True
    except Exception as error:
        print(f"No se pudo instalar: {error}")
        print("Ejecuta manualmente:")
        print("py -m pip install streamlit pandas numpy scipy statsmodels plotly openpyxl")
        return False


def mensaje_dependencias(faltantes: list[str]) -> str:
    lineas = [
        "Faltan librerías para ejecutar la aplicación.",
        "",
        "Librerías faltantes:",
        *[f"- {p}" for p in faltantes],
        "",
        "Solución:",
        "py -m pip install streamlit pandas numpy scipy statsmodels plotly openpyxl",
        "",
        "Luego ejecuta:",
        "py -m streamlit run PulsoCalidadSPC_FINAL_DEPURADA.py",
    ]
    return "\n".join(lineas)


if "--install" in sys.argv:
    instalar_paquetes(paquetes_faltantes(PAQUETES_APP))
    raise SystemExit(0)

FALTAN_BASE = paquetes_faltantes(PAQUETES_BASE)
FALTAN_APP = paquetes_faltantes(PAQUETES_APP)

if FALTAN_BASE:
    print(mensaje_dependencias(FALTAN_BASE))
else:
    import pandas as pd
    import numpy as np
    from scipy import stats
    from statsmodels.stats.stattools import durbin_watson


# ============================================================
# UTILIDADES GENERALES
# ============================================================

def verificar_dependencias_base() -> None:
    if FALTAN_BASE:
        raise RuntimeError(mensaje_dependencias(FALTAN_BASE))


def limpiar_nombre_columna(nombre) -> str:
    nombre = str(nombre).strip()
    reemplazos = {
        "LSL": "LIE",
        "USL": "LSE",
        "xbar": "Xbarra",
        "x-bar": "Xbarra",
    }
    return reemplazos.get(nombre, nombre)


def normalizar_nombres_columnas(df):
    verificar_dependencias_base()
    df = df.copy()
    nuevas = []
    usados = {}
    for col in df.columns:
        nombre = limpiar_nombre_columna(col)
        if nombre in usados:
            usados[nombre] += 1
            nombre = f"{nombre}_{usados[nombre]}"
        else:
            usados[nombre] = 0
        nuevas.append(nombre)
    df.columns = nuevas
    return df


def dataframe_valido(df) -> bool:
    verificar_dependencias_base()
    return df is not None and isinstance(df, pd.DataFrame) and not df.empty


def obtener_columna(df, columna):
    verificar_dependencias_base()
    datos = df.loc[:, columna]
    if isinstance(datos, pd.DataFrame):
        datos = datos.iloc[:, 0]
    return datos


def convertir_a_numerica(serie):
    verificar_dependencias_base()
    if isinstance(serie, pd.DataFrame):
        serie = serie.iloc[:, 0]
    return pd.to_numeric(serie, errors="coerce").dropna()


def columnas_numericas(df) -> list[str]:
    verificar_dependencias_base()
    if not dataframe_valido(df):
        return []
    cols = []
    for col in df.columns:
        if len(convertir_a_numerica(obtener_columna(df, col))) > 0:
            cols.append(col)
    return cols


def fmt(valor, decimales: Optional[int] = None):
    if isinstance(valor, (float, int, np.integer, np.floating)) and not pd.isna(valor):
        valor = float(valor)
        if valor == 0:
            return "0"
        if abs(valor) >= 100000:
            return f"{valor:,.0f}"
        if abs(valor) >= 1000:
            return f"{valor:,.1f}".rstrip("0").rstrip(".")
        if abs(valor) >= 100:
            return f"{valor:,.2f}".rstrip("0").rstrip(".")
        if abs(valor) >= 10:
            return f"{valor:,.3f}".rstrip("0").rstrip(".")
        if abs(valor) >= 1:
            return f"{valor:,.4f}".rstrip("0").rstrip(".")
        if abs(valor) >= 0.01:
            return f"{valor:,.5f}".rstrip("0").rstrip(".")
        return f"{valor:,.8f}".rstrip("0").rstrip(".")
    return str(valor)


def valor_limpio(valor):
    if isinstance(valor, (float, int, np.integer, np.floating)) and not pd.isna(valor):
        texto = fmt(valor)
        try:
            return float(texto.replace(",", ""))
        except Exception:
            return texto
    return valor


def redondear_dict(datos: dict, decimales: Optional[int] = None) -> dict:
    salida = {}
    for k, v in datos.items():
        salida[k] = valor_limpio(v)
    return salida


def guardar_estado_si_no_existe(st, nombre, valor):
    if nombre not in st.session_state:
        st.session_state[nombre] = valor


def _actualizar_estado_persistente(nombre):
    import streamlit as st
    clave_widget = f"_widget_{nombre}"
    if clave_widget in st.session_state:
        st.session_state[nombre] = st.session_state[clave_widget]


def numero_persistente(st, etiqueta, nombre, valor_defecto, **kwargs):
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state:
        st.session_state[clave_widget] = st.session_state[nombre]
    return st.number_input(
        etiqueta,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


# ============================================================
# PARÁMETROS COMPARTIDOS ENTRE MÓDULOS
# ============================================================

def _set_persistente(st, nombre, valor):
    """Actualiza un valor base sin modificar widgets ya creados.

    Streamlit no permite cambiar st.session_state de una llave usada por un
    widget después de que el widget fue instanciado en la misma ejecución. Por
    eso se actualiza siempre la llave lógica y solo se inicializa la llave del
    widget cuando todavía no existe.
    """
    try:
        if isinstance(valor, (np.integer, np.floating)):
            valor = float(valor)
    except Exception:
        pass
    st.session_state[nombre] = valor
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state:
        st.session_state[clave_widget] = valor


def parametros_proceso(st) -> dict:
    """Devuelve el banco central de parámetros técnicos del proceso."""
    return st.session_state.get("parametros_proceso", {}) or {}


def guardar_parametros_proceso(st, **kwargs) -> None:
    """Guarda parámetros que deben viajar entre capacidad, PNC, diseño, reporte y conclusiones."""
    params = parametros_proceso(st).copy()
    for k, v in kwargs.items():
        if v is not None:
            try:
                if isinstance(v, (np.integer, np.floating)):
                    v = float(v)
            except Exception:
                pass
            params[k] = v
    params["actualizado"] = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
    st.session_state["parametros_proceso"] = params


def sincronizar_parametros_widgets(st) -> None:
    """Envía los parámetros técnicos activos a los campos de los módulos siguientes."""
    p = parametros_proceso(st)
    if not p:
        return
    mapa = {
        "vn": ["cap_vn", "nc_vn", "rep_vn", "conc_vn"],
        "lie": ["cap_lie", "nc_lie", "rep_lie", "conc_lie"],
        "lse": ["cap_lse", "nc_lse", "rep_lse", "conc_lse"],
        "media": ["cap_media_conocida", "dis_media_actual"],
        "sigma": ["cap_sigma_historica", "dis_sigma", "rep_sigma_hist"],
        "nsub": ["cap_nsub", "dis_n"],
        "intervalo": ["cap_intervalo", "dis_intervalo"],
    }
    for origen, destinos in mapa.items():
        if origen in p:
            for destino in destinos:
                _set_persistente(st, destino, p[origen])
    if "variable" in p:
        for destino in ["cap_variable", "nc_variable", "sup_variable", "rep_variable", "conc_variable"]:
            st.session_state[destino] = p["variable"]
            clave_widget = f"_widget_{destino}"
            if clave_widget not in st.session_state:
                st.session_state[clave_widget] = p["variable"]
    if "media_detectar" in p:
        _set_persistente(st, "cap_media_detectar", p["media_detectar"])
        _set_persistente(st, "dis_media_cambio", p["media_detectar"])


def actualizar_parametros_desde_df(st, df, variable=None) -> None:
    """Crea parámetros lógicos iniciales al cargar o cambiar datos."""
    if not dataframe_valido(df):
        return
    nums = columnas_numericas(df)
    if not nums:
        return
    if variable not in nums:
        variable = nums[0]
    s = convertir_a_numerica(obtener_columna(df, variable))
    if len(s) == 0:
        return
    media = float(s.mean())
    sigma = float(s.std(ddof=1)) if len(s) > 1 else 1.0
    if pd.isna(sigma) or sigma <= 0:
        sigma = 1.0
    lie = float(s.min())
    lse = float(s.max())
    vn = float((lie + lse) / 2) if lie < lse else media
    tolerancia = float(max(abs(lse - vn), abs(vn - lie), sigma))
    guardar_parametros_proceso(
        st,
        variable=variable,
        media=media,
        sigma=sigma,
        vn=vn,
        lie=lie,
        lse=lse,
        tolerancia=tolerancia,
        n=int(len(s)),
    )
    sincronizar_parametros_widgets(st)


def panel_parametros_activos(st):
    """Muestra los parámetros que la app reutiliza en los módulos posteriores."""
    p = parametros_proceso(st)
    if not p:
        caja_estado(st, "info", "Cuando cargues datos o calcules capacidad, la app guardará media, sigma, VN, LIE y LSE para usarlos en los demás módulos.")
        return
    mostrar = {
        "Variable activa": p.get("variable", "No definida"),
        "Media usada": p.get("media", np.nan),
        "Sigma usada": p.get("sigma", np.nan),
        "VN": p.get("vn", np.nan),
        "LIE": p.get("lie", np.nan),
        "LSE": p.get("lse", np.nan),
    }
    tarjetas(st, mostrar, "Parámetros activos del proceso")
    caja_estado(st, "info", "Estos valores alimentan automáticamente Capacidad, PNC, Diseño de gráficos, Reportes y Conclusiones. Puedes cambiarlos manualmente si tu ejercicio entrega especificaciones distintas.")


def checkbox_persistente(st, etiqueta, nombre, valor_defecto=False, **kwargs):
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state:
        st.session_state[clave_widget] = st.session_state[nombre]
    return st.checkbox(
        etiqueta,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def selectbox_persistente(st, etiqueta, opciones, nombre, valor_defecto=None, **kwargs):
    if valor_defecto is None:
        valor_defecto = opciones[0]
    if valor_defecto not in opciones:
        valor_defecto = opciones[0]
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    if st.session_state[nombre] not in opciones:
        st.session_state[nombre] = opciones[0]
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state or st.session_state[clave_widget] not in opciones:
        st.session_state[clave_widget] = st.session_state[nombre]
    indice = opciones.index(st.session_state[clave_widget])
    return st.selectbox(
        etiqueta,
        opciones,
        index=indice,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def radio_persistente(st, etiqueta, opciones, nombre, valor_defecto=None, **kwargs):
    if valor_defecto is None:
        valor_defecto = opciones[0]
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    if st.session_state[nombre] not in opciones:
        st.session_state[nombre] = opciones[0]
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state or st.session_state[clave_widget] not in opciones:
        st.session_state[clave_widget] = st.session_state[nombre]
    indice = opciones.index(st.session_state[clave_widget])
    return st.radio(
        etiqueta,
        opciones,
        index=indice,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def cargar_datos_desde_archivo(archivo):
    verificar_dependencias_base()
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        df = pd.read_csv(archivo)
    elif nombre.endswith(".xlsx"):
        df = pd.read_excel(archivo)
    else:
        raise ValueError("Formato no soportado. Usa CSV o XLSX.")
    return normalizar_nombres_columnas(df)


def convertir_ancho_a_largo(df, columna_subgrupo):
    verificar_dependencias_base()
    if columna_subgrupo not in df.columns:
        return df
    df_largo = df.melt(id_vars=columna_subgrupo, var_name="medicion_id", value_name="medicion")
    df_largo = df_largo.rename(columns={columna_subgrupo: "subgrupo"})
    df_largo["medicion"] = pd.to_numeric(df_largo["medicion"], errors="coerce")
    df_largo = df_largo.dropna(subset=["medicion"])
    return normalizar_nombres_columnas(df_largo)


# ============================================================
# CÁLCULOS ESTADÍSTICOS
# ============================================================

def resumen_descriptivo(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) == 0:
        return None
    return {
        "n": int(len(s)),
        "Media": float(s.mean()),
        "Mediana": float(s.median()),
        "Desv. estándar": float(s.std(ddof=1)) if len(s) > 1 else np.nan,
        "Mínimo": float(s.min()),
        "Máximo": float(s.max()),
        "Rango": float(s.max() - s.min()),
        "Q1": float(s.quantile(0.25)),
        "Q3": float(s.quantile(0.75)),
    }


def detectar_atipicos_iqr(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) == 0:
        return np.nan, np.nan, s
    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    li = q1 - 1.5 * iqr
    ls = q3 + 1.5 * iqr
    return float(li), float(ls), s[(s < li) | (s > ls)]


def evaluar_normalidad(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) < 3:
        return pd.DataFrame([{"Prueba": "Normalidad", "p-valor": np.nan, "Resultado": "Datos insuficientes", "Interpretación": "Se necesitan al menos 3 datos."}])
    pruebas = [("Shapiro-Wilk", stats.shapiro)]
    if len(s) >= 8:
        pruebas.append(("D'Agostino-Pearson", stats.normaltest))
    pruebas.append(("Jarque-Bera", stats.jarque_bera))
    filas = []
    for nombre, funcion in pruebas:
        try:
            est, p = funcion(s)
            resultado = "Cumple" if p >= 0.05 else "No cumple"
            filas.append({
                "Prueba": nombre,
                "Estadístico": round(float(est), 3),
                "p-valor": round(float(p), 3),
                "Resultado": resultado,
                "Interpretación": "Compatible con normalidad." if resultado == "Cumple" else "No se ajusta claramente a normalidad.",
            })
        except Exception as e:
            filas.append({"Prueba": nombre, "Estadístico": np.nan, "p-valor": np.nan, "Resultado": "No evaluado", "Interpretación": str(e)})
    return pd.DataFrame(filas)


def diagnostico_normalidad(tabla):
    if tabla is None or tabla.empty:
        return "No evaluado", "No hay resultados de normalidad."
    pvals = pd.to_numeric(tabla["p-valor"], errors="coerce").dropna()
    if len(pvals) == 0:
        return "No evaluado", "No hay p-valores suficientes."
    if (pvals >= 0.05).mean() >= 0.5:
        return "Cumple", "La mayoría de pruebas no rechaza normalidad con α = 0.05."
    return "No cumple", "La mayoría de pruebas rechaza normalidad con α = 0.05."


def evaluar_independencia(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie).reset_index(drop=True)
    if len(s) < 4:
        return {"Durbin-Watson": np.nan, "Autocorr. lag 1": np.nan, "Resultado": "Datos insuficientes", "Interpretación": "Se necesitan al menos 4 datos."}
    try:
        dw = float(durbin_watson(s))
        ac = float(s.autocorr(lag=1))
    except Exception as e:
        return {"Durbin-Watson": np.nan, "Autocorr. lag 1": np.nan, "Resultado": "No evaluado", "Interpretación": str(e)}
    cumple = 1.5 <= dw <= 2.5 and (not pd.isna(ac)) and abs(ac) <= 0.30
    return {
        "Durbin-Watson": round(dw, 3),
        "Autocorr. lag 1": round(ac, 3) if not pd.isna(ac) else np.nan,
        "Resultado": "Cumple" if cumple else "No cumple",
        "Interpretación": "No se detecta autocorrelación fuerte." if cumple else "Existe posible dependencia temporal, racha, tendencia o ciclo.",
    }


def evaluar_homocedasticidad(df, columna_valor, columna_grupo):
    verificar_dependencias_base()
    if not dataframe_valido(df) or columna_valor not in df.columns or columna_grupo not in df.columns:
        return pd.DataFrame([{"Prueba": "Homocedasticidad", "Resultado": "Columnas no válidas"}])
    grupos = []
    for _, g in df.groupby(columna_grupo):
        vals = convertir_a_numerica(obtener_columna(g, columna_valor))
        if len(vals) >= 2:
            grupos.append(vals)
    if len(grupos) < 2:
        return pd.DataFrame([{"Prueba": "Homocedasticidad", "Resultado": "Datos insuficientes", "Interpretación": "Se requieren al menos 2 grupos con 2 datos."}])
    filas = []
    for nombre, funcion in [("Levene", stats.levene), ("Bartlett", stats.bartlett)]:
        try:
            est, p = funcion(*grupos)
            resultado = "Cumple" if p >= 0.05 else "No cumple"
            filas.append({
                "Prueba": nombre,
                "Estadístico": round(float(est), 3),
                "p-valor": round(float(p), 3),
                "Resultado": resultado,
                "Interpretación": "Varianzas similares." if resultado == "Cumple" else "Varianzas diferentes.",
            })
        except Exception as e:
            filas.append({"Prueba": nombre, "Resultado": "No evaluado", "Interpretación": str(e)})
    return pd.DataFrame(filas)


# ============================================================
# CAPACIDAD
# ============================================================

def clasificar_capacidad(cpk: float) -> str:
    if cpk >= 1.67:
        return "Excelente"
    if cpk >= 1.33:
        return "Capaz"
    if cpk >= 1.00:
        return "Marginal"
    return "No capaz"


def calcular_capacidad(serie, lie, lse, vn=None, sigma_historica=None, media_conocida=None):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) < 2 or lie >= lse:
        return None
    media_muestral = float(s.mean())
    media = float(media_conocida) if media_conocida is not None else media_muestral
    sigma_m = float(s.std(ddof=1))
    sigma = float(sigma_historica) if sigma_historica is not None and sigma_historica > 0 else sigma_m
    if sigma <= 0 or pd.isna(sigma):
        return None
    if vn is None:
        vn = (lie + lse) / 2
    cp = (lse - lie) / (6 * sigma)
    cpu = (lse - media) / (3 * sigma)
    cpl = (media - lie) / (3 * sigma)
    cpk = min(cpu, cpl)
    pp = (lse - lie) / (6 * sigma_m)
    ppu = (lse - media) / (3 * sigma_m)
    ppl = (media - lie) / (3 * sigma_m)
    ppk = min(ppu, ppl)
    cpm = (lse - lie) / (6 * np.sqrt(sigma**2 + (media - vn) ** 2))
    obs_lie = int((s < lie).sum())
    obs_lse = int((s > lse).sum())
    obs_total = obs_lie + obs_lse
    p_lie = float(stats.norm.cdf(lie, loc=media, scale=sigma))
    p_lse = float(1 - stats.norm.cdf(lse, loc=media, scale=sigma))
    p_total = p_lie + p_lse
    return {
        "n": int(len(s)),
        "VN": float(vn),
        "LIE": float(lie),
        "LSE": float(lse),
        "Media usada": media,
        "Media muestral": media_muestral,
        "Media conocida usada": "Sí" if media_conocida is not None else "No",
        "Sigma usada": sigma,
        "Sigma muestral": sigma_m,
        "Sigma histórica usada": "Sí" if sigma_historica is not None and sigma_historica > 0 else "No",
        "Cp": float(cp),
        "CPU": float(cpu),
        "CPL": float(cpl),
        "Cpk": float(cpk),
        "Pp": float(pp),
        "PPU": float(ppu),
        "PPL": float(ppl),
        "Ppk": float(ppk),
        "Cpm": float(cpm),
        "Z LIE": float((media - lie) / sigma),
        "Z LSE": float((lse - media) / sigma),
        "Producto no conforme observado": int(obs_total),
        "Fuera por LIE observado": int(obs_lie),
        "Fuera por LSE observado": int(obs_lse),
        "% PNC observado": float(obs_total / len(s) * 100),
        "% PNC estimado LIE": p_lie * 100,
        "% PNC estimado LSE": p_lse * 100,
        "% PNC estimado total": p_total * 100,
        "PPM estimado": p_total * 1_000_000,
        "Estado": clasificar_capacidad(cpk),
        "Centrado": "Centrado" if abs(media - vn) <= 0.25 * sigma else "Descentrado",
        "Riesgo principal": "LSE" if cpu < cpl else "LIE" if cpl < cpu else "Equilibrado",
    }


def diagnostico_capacidad(res):
    if res is None:
        return ["No se pudo calcular capacidad. Revisa que existan datos numéricos, que LIE sea menor que LSE y que la desviación estándar sea mayor que cero."]

    mensajes = []
    cpk = res["Cpk"]
    cp = res["Cp"]
    cpl = res["CPL"]
    cpu = res["CPU"]
    pnc_total = res["% PNC estimado total"]
    pnc_obs = res["% PNC observado"]

    mensajes.append(f"Conclusión general: el proceso se clasifica como {res['Estado']}. Esta clasificación se basa principalmente en Cpk, porque Cpk considera al mismo tiempo la variabilidad y el centrado frente a los límites de especificación.")

    if cp < 1:
        mensajes.append("Cp es menor que 1. Esto indica que la variabilidad natural del proceso es más ancha que la tolerancia disponible entre LIE y LSE. Aunque el proceso estuviera centrado, seguiría teniendo riesgo alto de generar producto fuera de especificación.")
    elif cp < 1.33:
        mensajes.append("Cp está entre 1 y 1.33. La variabilidad cabe dentro de la tolerancia, pero el margen es limitado. El proceso requiere control cuidadoso y reducción de variación para trabajar con seguridad.")
    else:
        mensajes.append("Cp es igual o mayor que 1.33. La variabilidad potencial del proceso es adecuada frente a la tolerancia, siempre que el proceso esté estable y bien centrado.")

    if cpk < cp * 0.85:
        mensajes.append("Cpk es bastante menor que Cp. Esto evidencia descentrado: el problema no es solo la variabilidad, sino también la ubicación de la media dentro de los límites de especificación.")
    else:
        mensajes.append("Cpk es cercano a Cp. Esto indica que el proceso está relativamente centrado respecto a sus límites de especificación.")

    if cpu < cpl:
        mensajes.append("CPU es menor que CPL. El lado crítico es el límite superior de especificación LSE. La media está más cerca del límite superior, por eso el mayor riesgo es producir valores altos fuera de especificación.")
    elif cpl < cpu:
        mensajes.append("CPL es menor que CPU. El lado crítico es el límite inferior de especificación LIE. La media está más cerca del límite inferior, por eso el mayor riesgo es producir valores bajos fuera de especificación.")
    else:
        mensajes.append("CPU y CPL son similares. El riesgo está equilibrado entre el límite inferior y el límite superior.")

    if res["Centrado"] == "Descentrado":
        mensajes.append("El proceso está descentrado respecto al valor nominal. Para mejorar Cpk no basta con reducir variación; también se debe ajustar la media del proceso hacia el objetivo o hacia el centro de especificación.")
    else:
        mensajes.append("El proceso aparece centrado respecto al valor nominal. En este caso, la mejora principal debe enfocarse en reducir variabilidad si Cp o Cpk son bajos.")

    if pnc_total > 5:
        mensajes.append(f"El producto no conforme estimado es alto: {fmt(pnc_total)}%. Este nivel indica pérdidas relevantes, reproceso o incumplimiento frecuente de especificaciones.")
    elif pnc_total > 1:
        mensajes.append(f"El producto no conforme estimado es moderado: {fmt(pnc_total)}%. Conviene mejorar el proceso antes de usarlo como proceso estable de producción.")
    else:
        mensajes.append(f"El producto no conforme estimado es bajo: {fmt(pnc_total)}%. Aun así, debe verificarse estabilidad con cartas de control antes de concluir capacidad final.")

    if abs(pnc_obs - pnc_total) <= max(1, pnc_total * 0.25):
        mensajes.append("El PNC observado y el PNC estimado son cercanos. Esto sugiere coherencia entre los datos observados y el modelo normal usado en capacidad.")
    else:
        mensajes.append("El PNC observado y el PNC estimado difieren de forma importante. Revisa normalidad, datos atípicos, estabilidad del proceso o si la distribución real no es normal.")

    if res["Producto no conforme observado"] > 0:
        mensajes.append(f"En los datos se observaron {res['Producto no conforme observado']} unidades fuera de especificación: {res['Fuera por LIE observado']} por debajo de LIE y {res['Fuera por LSE observado']} por encima de LSE.")
    else:
        mensajes.append("En la muestra no se observaron unidades fuera de especificación. Esto no garantiza capacidad si Cp o Cpk son bajos; solo indica que la muestra analizada no presentó fallos observados.")

    mensajes.append("Recomendación técnica: primero confirma estabilidad con la carta de control correspondiente. Después ajusta el centrado de la media y reduce la variabilidad. La capacidad solo debe interpretarse como válida cuando el proceso está bajo control estadístico.")
    return mensajes


def calcular_sigma_objetivo(lie, lse, cp_objetivo):
    if cp_objetivo <= 0:
        return np.nan
    return float((lse - lie) / (6 * cp_objetivo))


def media_maxima_para_pnc_lse(lse, sigma, pnc_lse):
    z = stats.norm.ppf(1 - pnc_lse)
    return float(lse - z * sigma)


def potencia_xbarra(media_actual, media_cambio, sigma, n, z=3):
    if sigma <= 0 or n <= 0:
        return None
    diferencia = abs(media_cambio - media_actual)
    d = diferencia / sigma
    d_raiz_n = d * np.sqrt(n)
    z_menos = z - d_raiz_n
    beta = float(stats.norm.cdf(z_menos))
    potencia = float(1 - beta)
    arl = np.inf if potencia <= 0 else 1 / potencia
    return {
        "Media actual": float(media_actual),
        "Media a detectar": float(media_cambio),
        "Sigma usada": float(sigma),
        "Diferencia absoluta": float(diferencia),
        "d": float(d),
        "d√n": float(d_raiz_n),
        "Z crítico": float(z),
        "Z crítico - d√n": float(z_menos),
        "β": float(beta),
        "% β": float(beta * 100),
        "Potencia = 1 - β": potencia,
        "% Potencia": potencia * 100,
        "ARL1": arl,
    }


def n_para_potencia(media_actual, media_cambio, sigma, potencia_objetivo, z=3):
    if sigma <= 0 or potencia_objetivo <= 0 or potencia_objetivo >= 1 or media_actual == media_cambio:
        return None
    d = abs(media_cambio - media_actual) / sigma
    z_beta = stats.norm.ppf(1 - potencia_objetivo)
    return int(np.ceil(((z - z_beta) / d) ** 2))


# ============================================================
# CARTAS DE CONTROL
# ============================================================

def constantes_xbar_r(n):
    tabla = {
        2: {"A2": 1.880, "D3": 0.000, "D4": 3.267, "d2": 1.128},
        3: {"A2": 1.023, "D3": 0.000, "D4": 2.574, "d2": 1.693},
        4: {"A2": 0.729, "D3": 0.000, "D4": 2.282, "d2": 2.059},
        5: {"A2": 0.577, "D3": 0.000, "D4": 2.114, "d2": 2.326},
        6: {"A2": 0.483, "D3": 0.000, "D4": 2.004, "d2": 2.534},
        7: {"A2": 0.419, "D3": 0.076, "D4": 1.924, "d2": 2.704},
        8: {"A2": 0.373, "D3": 0.136, "D4": 1.864, "d2": 2.847},
        9: {"A2": 0.337, "D3": 0.184, "D4": 1.816, "d2": 2.970},
        10: {"A2": 0.308, "D3": 0.223, "D4": 1.777, "d2": 3.078},
    }
    return tabla.get(int(n))


def calcular_xbar_r(df, col_valor, col_subgrupo):
    if not dataframe_valido(df) or col_valor not in df.columns or col_subgrupo not in df.columns:
        return None
    temp = pd.DataFrame({
        "valor": pd.to_numeric(obtener_columna(df, col_valor), errors="coerce"),
        "subgrupo": obtener_columna(df, col_subgrupo),
    }).dropna()
    if temp.empty:
        return None
    resumen = temp.groupby("subgrupo")["valor"].agg(["mean", "max", "min", "count"])
    resumen["R"] = resumen["max"] - resumen["min"]
    resumen = resumen[resumen["count"] >= 2]
    if resumen.empty:
        return None
    n = int(round(float(resumen["count"].mean())))
    c = constantes_xbar_r(n)
    if not c:
        return None
    xbb = float(resumen["mean"].mean())
    rb = float(resumen["R"].mean())
    limites = {
        "Tamaño promedio de subgrupo": n,
        "X doble barra": xbb,
        "R barra": rb,
        "A2": c["A2"],
        "D3": c["D3"],
        "D4": c["D4"],
        "d2": c["d2"],
        "LC Xbarra": xbb,
        "LCS Xbarra": xbb + c["A2"] * rb,
        "LCI Xbarra": xbb - c["A2"] * rb,
        "LC R": rb,
        "LCS R": c["D4"] * rb,
        "LCI R": c["D3"] * rb,
        "Sigma Rbarra/d2": rb / c["d2"],
    }
    limites["Puntos fuera Xbarra"] = int(((resumen["mean"] > limites["LCS Xbarra"]) | (resumen["mean"] < limites["LCI Xbarra"])).sum())
    limites["Puntos fuera R"] = int(((resumen["R"] > limites["LCS R"]) | (resumen["R"] < limites["LCI R"])).sum())
    return resumen, limites


def calcular_i_mr(serie):
    s = convertir_a_numerica(serie).reset_index(drop=True)
    if len(s) < 3:
        return None
    mr = s.diff().abs().dropna()
    mr_barra = float(mr.mean())
    if mr_barra == 0 or pd.isna(mr_barra):
        return None
    media = float(s.mean())
    limites = {
        "LC I": media,
        "LCS I": media + 2.66 * mr_barra,
        "LCI I": media - 2.66 * mr_barra,
        "MR barra": mr_barra,
        "LC MR": mr_barra,
        "LCS MR": 3.268 * mr_barra,
        "LCI MR": 0.0,
    }
    limites["Puntos fuera I"] = int(((s > limites["LCS I"]) | (s < limites["LCI I"])).sum())
    limites["Puntos fuera MR"] = int(((mr > limites["LCS MR"]) | (mr < limites["LCI MR"])).sum())
    return s, mr, limites


def reglas_shewhart(valores, lc, lcs, lci):
    s = convertir_a_numerica(valores).reset_index(drop=True)
    if len(s) == 0:
        return pd.DataFrame(columns=["Regla", "Resultado"])
    fuera = int(((s > lcs) | (s < lci)).sum())
    arriba = (s > lc).astype(int)
    abajo = (s < lc).astype(int)
    max_racha = 0
    actual = 0
    lado_anterior = None
    for a, b in zip(arriba, abajo):
        lado = "arriba" if a else "abajo" if b else "centro"
        if lado != "centro" and lado == lado_anterior:
            actual += 1
        else:
            actual = 1
            lado_anterior = lado
        max_racha = max(max_racha, actual)
    tendencias = 0
    dif = s.diff().dropna()
    if len(dif) >= 5:
        for i in range(len(dif) - 4):
            v = dif.iloc[i:i + 5]
            if (v > 0).all() or (v < 0).all():
                tendencias += 1
    return pd.DataFrame([
        {"Regla": "Puntos fuera de límites", "Resultado": fuera},
        {"Regla": "Mayor racha a un lado de LC", "Resultado": int(max_racha)},
        {"Regla": "Tendencias de 6 puntos", "Resultado": int(tendencias)},
    ])


def calcular_p(df, col_def, col_n):
    temp = pd.DataFrame({"def": pd.to_numeric(obtener_columna(df, col_def), errors="coerce"), "n": pd.to_numeric(obtener_columna(df, col_n), errors="coerce")}).dropna()
    temp = temp[(temp["n"] > 0) & (temp["def"] >= 0) & (temp["def"] <= temp["n"])]
    if temp.empty:
        return None
    pbar = float(temp["def"].sum() / temp["n"].sum())
    temp["p"] = temp["def"] / temp["n"]
    temp["LCS"] = (pbar + 3 * np.sqrt(pbar * (1 - pbar) / temp["n"])).clip(upper=1)
    temp["LCI"] = (pbar - 3 * np.sqrt(pbar * (1 - pbar) / temp["n"])).clip(lower=0)
    limites = {"p barra": pbar, "Defectuosos totales": int(temp["def"].sum()), "Inspeccionados totales": int(temp["n"].sum()), "% no conforme": pbar * 100}
    limites["Puntos fuera"] = int(((temp["p"] > temp["LCS"]) | (temp["p"] < temp["LCI"])).sum())
    return temp, limites


def calcular_np(df, col_def, col_n):
    temp = pd.DataFrame({"def": pd.to_numeric(obtener_columna(df, col_def), errors="coerce"), "n": pd.to_numeric(obtener_columna(df, col_n), errors="coerce")}).dropna()
    temp = temp[(temp["n"] > 0) & (temp["def"] >= 0) & (temp["def"] <= temp["n"])]
    if temp.empty:
        return None
    n_usado = float(temp["n"].mean())
    pbar = float(temp["def"].sum() / temp["n"].sum())
    npbar = n_usado * pbar
    sigma = np.sqrt(n_usado * pbar * (1 - pbar))
    limites = {"n usado": n_usado, "p barra": pbar, "LC np": npbar, "LCS np": npbar + 3 * sigma, "LCI np": max(0, npbar - 3 * sigma)}
    limites["Puntos fuera"] = int(((temp["def"] > limites["LCS np"]) | (temp["def"] < limites["LCI np"])).sum())
    return temp, limites


def calcular_c(df, col_defectos):
    temp = pd.DataFrame({"c": pd.to_numeric(obtener_columna(df, col_defectos), errors="coerce")}).dropna()
    temp = temp[temp["c"] >= 0]
    if temp.empty:
        return None
    cbar = float(temp["c"].mean())
    limites = {"LC c": cbar, "LCS c": cbar + 3 * np.sqrt(cbar), "LCI c": max(0, cbar - 3 * np.sqrt(cbar))}
    limites["Puntos fuera"] = int(((temp["c"] > limites["LCS c"]) | (temp["c"] < limites["LCI c"])).sum())
    return temp, limites


def calcular_u(df, col_defectos, col_unidades):
    temp = pd.DataFrame({"def": pd.to_numeric(obtener_columna(df, col_defectos), errors="coerce"), "n": pd.to_numeric(obtener_columna(df, col_unidades), errors="coerce")}).dropna()
    temp = temp[(temp["def"] >= 0) & (temp["n"] > 0)]
    if temp.empty:
        return None
    ubar = float(temp["def"].sum() / temp["n"].sum())
    temp["u"] = temp["def"] / temp["n"]
    temp["LCS"] = ubar + 3 * np.sqrt(ubar / temp["n"])
    temp["LCI"] = (ubar - 3 * np.sqrt(ubar / temp["n"])).clip(lower=0)
    limites = {"u barra": ubar, "Defectos totales": int(temp["def"].sum()), "Unidades totales": int(temp["n"].sum())}
    limites["Puntos fuera"] = int(((temp["u"] > temp["LCS"]) | (temp["u"] < temp["LCI"])).sum())
    return temp, limites


def curva_oc(n, c, pasos=151):
    ps = np.linspace(0, 0.25, pasos)
    return pd.DataFrame({"Fracción defectuosa": ps, "Probabilidad de aceptación": [float(stats.binom.cdf(min(c, n), n, p)) for p in ps]})




def constantes_xbar_s(n):
    tabla = {
        2: {"A3": 2.659, "B3": 0.000, "B4": 3.267, "c4": 0.7979},
        3: {"A3": 1.954, "B3": 0.000, "B4": 2.568, "c4": 0.8862},
        4: {"A3": 1.628, "B3": 0.000, "B4": 2.266, "c4": 0.9213},
        5: {"A3": 1.427, "B3": 0.000, "B4": 2.089, "c4": 0.9400},
        6: {"A3": 1.287, "B3": 0.030, "B4": 1.970, "c4": 0.9515},
        7: {"A3": 1.182, "B3": 0.118, "B4": 1.882, "c4": 0.9594},
        8: {"A3": 1.099, "B3": 0.185, "B4": 1.815, "c4": 0.9650},
        9: {"A3": 1.032, "B3": 0.239, "B4": 1.761, "c4": 0.9693},
        10: {"A3": 0.975, "B3": 0.284, "B4": 1.716, "c4": 0.9727},
        11: {"A3": 0.927, "B3": 0.321, "B4": 1.679, "c4": 0.9754},
        12: {"A3": 0.886, "B3": 0.354, "B4": 1.646, "c4": 0.9776},
        13: {"A3": 0.850, "B3": 0.382, "B4": 1.618, "c4": 0.9794},
        14: {"A3": 0.817, "B3": 0.406, "B4": 1.594, "c4": 0.9810},
        15: {"A3": 0.789, "B3": 0.428, "B4": 1.572, "c4": 0.9823},
        20: {"A3": 0.680, "B3": 0.510, "B4": 1.490, "c4": 0.9869},
        25: {"A3": 0.606, "B3": 0.565, "B4": 1.435, "c4": 0.9896},
    }
    return tabla.get(int(n))


def calcular_xbar_s(df, col_valor, col_subgrupo):
    if not dataframe_valido(df) or col_valor not in df.columns or col_subgrupo not in df.columns:
        return None
    temp = pd.DataFrame({
        "valor": pd.to_numeric(obtener_columna(df, col_valor), errors="coerce"),
        "subgrupo": obtener_columna(df, col_subgrupo),
    }).dropna()
    if temp.empty:
        return None
    resumen = temp.groupby("subgrupo")["valor"].agg(["mean", "std", "count"])
    resumen = resumen[(resumen["count"] >= 2) & resumen["std"].notna()]
    if resumen.empty:
        return None
    n = int(round(float(resumen["count"].mean())))
    c = constantes_xbar_s(n)
    if not c:
        return None
    xbb = float(resumen["mean"].mean())
    sb = float(resumen["std"].mean())
    limites = {
        "Tamaño promedio de subgrupo": n,
        "X doble barra": xbb,
        "S barra": sb,
        "A3": c["A3"],
        "B3": c["B3"],
        "B4": c["B4"],
        "c4": c["c4"],
        "LC Xbarra": xbb,
        "LCS Xbarra": xbb + c["A3"] * sb,
        "LCI Xbarra": xbb - c["A3"] * sb,
        "LC S": sb,
        "LCS S": c["B4"] * sb,
        "LCI S": c["B3"] * sb,
        "Sigma Sbarra/c4": sb / c["c4"],
    }
    limites["Puntos fuera Xbarra"] = int(((resumen["mean"] > limites["LCS Xbarra"]) | (resumen["mean"] < limites["LCI Xbarra"])).sum())
    limites["Puntos fuera S"] = int(((resumen["std"] > limites["LCS S"]) | (resumen["std"] < limites["LCI S"])).sum())
    return resumen, limites


def evaluar_patrones_shewhart(valores, lc, lcs, lci):
    s = convertir_a_numerica(valores).reset_index(drop=True)
    if len(s) == 0:
        return pd.DataFrame(columns=["Regla", "Resultado", "Interpretación"])
    sigma = (lcs - lc) / 3 if lcs is not None and lc is not None else np.nan
    fuera = int(((s > lcs) | (s < lci)).sum())
    filas = [{"Regla": "1 punto fuera de LCI/LCS", "Resultado": fuera, "Interpretación": "Señal directa de causa especial." if fuera else "No se detecta esta señal."}]
    racha_max = 0
    racha = 0
    lado_prev = None
    for v in s:
        lado = "arriba" if v > lc else "abajo" if v < lc else "centro"
        if lado != "centro" and lado == lado_prev:
            racha += 1
        elif lado != "centro":
            racha = 1
            lado_prev = lado
        else:
            racha = 0
            lado_prev = None
        racha_max = max(racha_max, racha)
    filas.append({"Regla": "8 puntos seguidos a un lado de LC", "Resultado": int(racha_max >= 8), "Interpretación": f"Racha máxima observada: {racha_max}."})
    tendencia = 0
    if len(s) >= 6:
        for i in range(len(s) - 5):
            tramo = s.iloc[i:i+6]
            dif = tramo.diff().dropna()
            if (dif > 0).all() or (dif < 0).all():
                tendencia += 1
    filas.append({"Regla": "6 puntos consecutivos con tendencia", "Resultado": int(tendencia), "Interpretación": "Indica deriva gradual del proceso." if tendencia else "No se detecta tendencia larga."})
    dos_de_tres = 0
    if len(s) >= 3 and sigma > 0:
        for i in range(len(s) - 2):
            w = s.iloc[i:i+3]
            if ((w > lc + 2*sigma).sum() >= 2) or ((w < lc - 2*sigma).sum() >= 2):
                dos_de_tres += 1
    filas.append({"Regla": "2 de 3 puntos en zona A", "Resultado": int(dos_de_tres), "Interpretación": "Señal temprana de desplazamiento." if dos_de_tres else "No se detecta esta señal."})
    cuatro_de_cinco = 0
    if len(s) >= 5 and sigma > 0:
        for i in range(len(s) - 4):
            w = s.iloc[i:i+5]
            if ((w > lc + sigma).sum() >= 4) or ((w < lc - sigma).sum() >= 4):
                cuatro_de_cinco += 1
    filas.append({"Regla": "4 de 5 puntos en zona B o más", "Resultado": int(cuatro_de_cinco), "Interpretación": "Sugiere cambio sostenido." if cuatro_de_cinco else "No se detecta esta señal."})
    return pd.DataFrame(filas)


def calcular_desempeno_grafico_xbarra(media_actual, media_cambio, sigma, n, intervalo=1.0, z=3.0):
    if sigma <= 0 or n <= 0 or intervalo <= 0:
        return None
    d = abs(media_cambio - media_actual) / sigma
    alpha = 2 * (1 - stats.norm.cdf(abs(z)))
    beta = float(stats.norm.cdf(abs(z) - d * np.sqrt(n)) - stats.norm.cdf(-abs(z) - d * np.sqrt(n)))
    beta = min(max(beta, 0.0), 1.0)
    potencia = 1 - beta
    arl0 = np.inf if alpha <= 0 else 1 / alpha
    arl1 = np.inf if potencia <= 0 else 1 / potencia
    return {
        "media actual": float(media_actual), "media a detectar": float(media_cambio), "sigma usada": float(sigma), "n": int(n),
        "intervalo minutos": float(intervalo), "diferencia absoluta": float(abs(media_cambio - media_actual)), "d": float(d),
        "d√n": float(d * np.sqrt(n)), "Z crítico": float(abs(z)), "Z crítico - d√n": float(abs(z) - d * np.sqrt(n)),
        "α": float(alpha), "% α": float(alpha * 100), "β": float(beta), "% β": float(beta * 100),
        "potencia = 1 - β": float(potencia), "% potencia": float(potencia * 100),
        "ARL0": float(arl0) if np.isfinite(arl0) else np.inf, "ATS0 minutos": float(arl0 * intervalo) if np.isfinite(arl0) else np.inf,
        "ARL1": float(arl1) if np.isfinite(arl1) else np.inf, "ATS1 minutos": float(arl1 * intervalo) if np.isfinite(arl1) else np.inf,
    }


def n_para_potencia_xbarra(media_actual, media_cambio, sigma, potencia_objetivo, z=3.0):
    if sigma <= 0 or potencia_objetivo <= 0 or potencia_objetivo >= 1 or media_actual == media_cambio:
        return None
    d = abs(media_cambio - media_actual) / sigma
    z_beta = stats.norm.ppf(potencia_objetivo)
    return int(np.ceil(((abs(z) + z_beta) / d) ** 2))

# ============================================================
# INTERFAZ Y GRÁFICOS
# ============================================================

def importar_librerias_app():
    if FALTAN_APP:
        return None, None
    import streamlit as st
    import plotly.graph_objects as go
    return st, go


def aplicar_estilo_visual(st):
    st.markdown(
        """
<style>
@keyframes entradaSuave {from {opacity:0; transform:translateY(16px) scale(.985)} to {opacity:1; transform:translateY(0) scale(1)}}
@keyframes flotar {0%,100% {transform:translateY(0)} 50% {transform:translateY(-5px)}}
@keyframes brillo {0%,100% {box-shadow:0 12px 30px rgba(15,23,42,.08)} 50% {box-shadow:0 20px 48px rgba(14,165,233,.18)}}
@keyframes destello {0% {background-position:-360px 0} 100% {background-position:360px 0}}
@keyframes puntoVivo {0%,100% {transform:scale(1); opacity:.68} 50% {transform:scale(1.35); opacity:1}}
html, body, [class*="css"] {font-family:Arial, Helvetica, sans-serif;}
.stApp {
    background:
        radial-gradient(circle at 8% 4%, rgba(14,165,233,.20), transparent 30%),
        radial-gradient(circle at 94% 8%, rgba(168,85,247,.18), transparent 31%),
        radial-gradient(circle at 50% 100%, rgba(34,197,94,.12), transparent 30%),
        linear-gradient(135deg,#f8fafc 0%,#eef2ff 42%,#fdf4ff 72%,#f8fafc 100%);
}
.block-container {padding-top:1.20rem; padding-bottom:2.4rem; max-width:1380px; animation:entradaSuave .45s ease both;}
section[data-testid="stSidebar"] {background:linear-gradient(180deg,#111827 0%,#1e1b4b 54%,#0f172a 100%); border-right:1px solid rgba(255,255,255,.12);}
section[data-testid="stSidebar"] * {color:#f8fafc!important;}
div[role="radiogroup"] label {border-radius:20px!important; padding:.42rem .65rem!important; margin:.12rem 0!important; border:1px solid rgba(255,255,255,.08)!important; transition:all .18s ease!important;}
div[role="radiogroup"] label:hover {background:rgba(255,255,255,.13)!important; transform:translateX(5px) scale(1.01);}
.app-title {position:relative; overflow:hidden; padding:1.35rem 1.55rem; border-radius:34px; background:linear-gradient(135deg,rgba(255,255,255,.95),rgba(240,249,255,.90)); border:1px solid rgba(148,163,184,.25); box-shadow:0 20px 48px rgba(15,23,42,.10); margin-bottom:1.15rem; backdrop-filter:blur(14px); animation:entradaSuave .55s ease both;}
.app-title:before {content:""; position:absolute; width:190px; height:190px; right:-55px; top:-85px; background:linear-gradient(135deg,#38bdf8,#a78bfa,#fb7185); border-radius:50%; opacity:.18;}
.app-title:after {content:""; position:absolute; inset:0; background:linear-gradient(110deg,transparent,rgba(255,255,255,.58),transparent); background-size:360px 100%; animation:destello 4.6s linear infinite; pointer-events:none;}
.creator-tag {margin-top:.5rem; font-size:.78rem; color:#64748b; font-weight:600;}
.soft-card {min-height:108px; padding:1.05rem; border-radius:26px; background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(248,250,252,.91)); border:1px solid rgba(148,163,184,.22); box-shadow:0 14px 32px rgba(15,23,42,.08); transition:all .2s ease; animation:entradaSuave .45s ease both, flotar 6s ease-in-out infinite;}
.soft-card:hover {transform:translateY(-6px) scale(1.015); animation:brillo 1.7s ease infinite; border-color:rgba(14,165,233,.42);}
.mini-label {margin:0; color:#64748b; font-size:.73rem; letter-spacing:.055em; text-transform:uppercase; font-weight:800;}
.big-number {margin:.34rem 0 0 0; background:linear-gradient(90deg,#0f172a,#1d4ed8,#7c3aed); -webkit-background-clip:text; color:transparent; font-size:1.42rem; font-weight:900; line-height:1.1;}
.ok-box,.warn-box,.bad-box,.info-box {padding:1rem; border-radius:20px; margin:.45rem 0 .75rem 0; border:1px solid transparent; box-shadow:0 10px 24px rgba(15,23,42,.06); animation:entradaSuave .35s ease both;}
.ok-box {background:#ecfdf5; color:#065f46; border-color:#a7f3d0}.warn-box {background:#fffbeb; color:#92400e; border-color:#fde68a}.bad-box {background:#fef2f2; color:#991b1b; border-color:#fecaca}.info-box {background:#eff6ff; color:#1e3a8a; border-color:#bfdbfe}
div[data-testid="stDataFrame"] {border-radius:20px; overflow:hidden; box-shadow:0 16px 34px rgba(15,23,42,.08); animation:entradaSuave .45s ease both;}
.live-badge {display:inline-flex; gap:.4rem; padding:.48rem .8rem; border-radius:999px; background:#ecfeff; color:#155e75; border:1px solid #a5f3fc; font-weight:900; animation:brillo 1.6s ease-in-out infinite;}
.live-badge:before {content:'●'; color:#22c55e; animation:puntoVivo 1s ease-in-out infinite;}
button {border-radius:18px!important; transition:all .18s ease!important; font-weight:700!important;}
button:hover {transform:translateY(-2px)!important; box-shadow:0 12px 28px rgba(37,99,235,.15)!important;}
</style>
""",
        unsafe_allow_html=True,
    )


def encabezado(st, titulo, subtitulo):
    st.markdown(
        f"""
<div class="app-title">
    <h1 style="margin:0; color:#0f172a; font-size:2.05rem;">{titulo}</h1>
    <p style="margin:.25rem 0 0 0; color:#475569; font-size:1rem;">{subtitulo}</p>
    <p class="creator-tag">Creado por Jerson Andrés López Wilches</p>
</div>
""",
        unsafe_allow_html=True,
    )


def caja_estado(st, tipo, texto):
    clase = {"ok": "ok-box", "alerta": "warn-box", "error": "bad-box", "info": "info-box"}.get(tipo, "info-box")
    st.markdown(f"<div class='{clase}'>{texto}</div>", unsafe_allow_html=True)


def tarjetas(st, datos: dict, titulo: str):
    st.markdown(f"### {titulo}")
    claves = list(datos.keys())
    for i in range(0, len(claves), 4):
        cols = st.columns(min(4, len(claves) - i))
        for j, k in enumerate(claves[i:i + 4]):
            with cols[j]:
                st.markdown(f"<div class='soft-card'><p class='mini-label'>{k}</p><p class='big-number'>{fmt(datos[k])}</p></div>", unsafe_allow_html=True)


def tabla(st, datos: dict, titulo="Resultados"):
    st.markdown(f"### {titulo}")
    df = pd.DataFrame([redondear_dict(datos, 3)]).T.reset_index()
    df.columns = ["Indicador", "Valor"]
    st.dataframe(df, use_container_width=True, hide_index=True)



def excel_bytes_hojas(hojas: dict) -> bytes:
    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        for nombre, datos in hojas.items():
            if isinstance(datos, pd.DataFrame):
                df_export = datos.copy()
            elif isinstance(datos, dict):
                df_export = pd.DataFrame([redondear_dict(datos)]).T.reset_index()
                df_export.columns = ["Indicador", "Valor"]
            else:
                df_export = pd.DataFrame(datos)
            hoja = str(nombre)[:31] if nombre else "Hoja"
            df_export.to_excel(writer, index=False, sheet_name=hoja)
    return salida.getvalue()


def boton_exportar_excel(st, hojas: dict, nombre_archivo: str, etiqueta: str = "Descargar resultados en Excel"):
    try:
        st.download_button(
            etiqueta,
            data=excel_bytes_hojas(hojas),
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.warning(f"No se pudo preparar el Excel: {e}")


def perfil_datos(df):
    filas = []
    if not dataframe_valido(df):
        return pd.DataFrame()
    for col in df.columns:
        serie = df[col]
        num = pd.to_numeric(serie, errors="coerce")
        n_num = int(num.notna().sum())
        filas.append({
            "Variable": col,
            "Tipo detectado": "Numérica" if n_num > 0 else "Texto/Categoría",
            "Datos válidos": int(serie.notna().sum()),
            "Faltantes": int(serie.isna().sum()),
            "% faltantes": round(float(serie.isna().mean() * 100), 3),
            "Valores únicos": int(serie.nunique(dropna=True)),
            "Media si numérica": float(num.mean()) if n_num else np.nan,
            "Desv. estándar si numérica": float(num.std(ddof=1)) if n_num > 1 else np.nan,
        })
    return pd.DataFrame(filas)


def plan_modulos_df():
    return pd.DataFrame([
        {"Sección": "Gestión de datos", "Incluye": "Excel/CSV, identificación de variables, subgrupos racionales, limpieza y tablas resumen", "Estado": "Integrado"},
        {"Sección": "Gráficos de control por variables", "Incluye": "X-barra-R, X-barra-S e I-MR con Fase I y Fase II", "Estado": "Integrado"},
        {"Sección": "Gráficos de control por atributos", "Incluye": "p, np, c y u con selección según defecto, no conformidad y tamaño muestral", "Estado": "Integrado"},
        {"Sección": "Validación de supuestos", "Incluye": "Normalidad, independencia, estabilidad, patrones y causas especiales", "Estado": "Integrado"},
        {"Sección": "Capacidad del proceso", "Incluye": "Cp, Cpk, Cpl, Cpu, PNC, especificaciones y aptitud", "Estado": "Integrado"},
        {"Sección": "Diseño de gráficos", "Incluye": "ARL0, ARL1, ATS, tamaño de muestra, frecuencia y desplazamiento", "Estado": "Integrado"},
        {"Sección": "Muestreo de aceptación", "Incluye": "Diseño y validación de planes de muestreo por atributos", "Estado": "Integrado"},
        {"Sección": "Reportes", "Incluye": "Tablas, gráficos, conclusiones preliminares y exportación de resultados", "Estado": "Integrado"},
    ])


def texto_decision_empresarial(estado, riesgo, pnc_total=None):
    if estado == "Excelente" or estado == "Capaz":
        return "El proceso presenta condiciones favorables para producción continua. Se recomienda mantener seguimiento periódico y conservar los parámetros actuales."
    if pnc_total is not None and pnc_total > 5:
        return "El proceso requiere intervención prioritaria porque el porcentaje de producto no conforme es alto. Deben revisarse calibración, materia prima, método de operación y variabilidad."
    return f"El proceso requiere seguimiento técnico. El riesgo principal identificado es {riesgo}; se recomienda ajustar el centrado y controlar la variabilidad antes de liberar producción crítica."

def estilo_fig(fig, titulo, subtitulo=""):
    texto = titulo if not subtitulo else f"{titulo}<br><sup>{subtitulo}</sup>"
    fig.update_layout(
        title={"text": texto, "x": 0.03, "xanchor": "left", "font": {"size": 22}},
        template="plotly_white",
        height=510,
        margin=dict(l=48, r=34, t=90, b=52),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
        font=dict(size=13, family="Arial"),
        paper_bgcolor="rgba(255,255,255,0)",
        plot_bgcolor="rgba(255,255,255,.98)",
    )
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor="rgba(148,163,184,.20)", zeroline=False, showline=True, linecolor="rgba(15,23,42,.35)")
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor="rgba(148,163,184,.20)", zeroline=False, showline=True, linecolor="rgba(15,23,42,.35)")
    return fig


def hline(fig, y, nombre, dash="dash"):
    fig.add_hline(y=y, line_dash=dash, line_width=2.8, annotation_text=nombre, annotation_position="top left")


def vline(fig, x, nombre, dash="dash"):
    fig.add_vline(x=x, line_dash=dash, line_width=2.8, annotation_text=nombre, annotation_position="top")


def grafico_corrida(go, serie, titulo="Gráfico de corrida"):
    s = convertir_a_numerica(serie).reset_index(drop=True)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=s.index + 1, y=s, mode="lines+markers", name="Medición", line=dict(width=3), marker=dict(size=8)))
    if len(s):
        hline(fig, float(s.mean()), "Media", "solid")
    fig.update_xaxes(title="Orden")
    fig.update_yaxes(title="Medición")
    return estilo_fig(fig, titulo, "Evalúa tendencia, saltos, ciclos y comportamiento temporal")


def grafico_qq(go, serie):
    s = convertir_a_numerica(serie)
    fig = go.Figure()
    if len(s) < 3:
        return estilo_fig(fig, "Q-Q normal", "Datos insuficientes")
    osm, osr = stats.probplot(s, dist="norm", fit=False)
    pendiente, intercepto, _ = stats.probplot(s, dist="norm", fit=True)[1]
    xline = np.array([min(osm), max(osm)])
    yline = intercepto + pendiente * xline
    fig.add_trace(go.Scatter(x=osm, y=osr, mode="markers", name="Datos", marker=dict(size=9)))
    fig.add_trace(go.Scatter(x=xline, y=yline, mode="lines", name="Referencia normal", line=dict(width=4)))
    fig.update_xaxes(title="Cuantiles teóricos")
    fig.update_yaxes(title="Cuantiles observados")
    return estilo_fig(fig, "Gráfico Q-Q de normalidad", "Puntos cercanos a la línea indican ajuste normal")


def grafico_box(go, serie):
    s = convertir_a_numerica(serie)
    fig = go.Figure()
    fig.add_trace(go.Box(y=s, boxpoints="all", jitter=0.28, pointpos=0, name="Datos", marker=dict(size=7)))
    fig.update_yaxes(title="Medición")
    return estilo_fig(fig, "Boxplot de mediciones", "Visualiza dispersión, mediana y valores atípicos")


def grafico_hist_capacidad(go, serie, lie, lse, vn=None):
    s = convertir_a_numerica(serie)
    media = float(s.mean())
    sigma = float(s.std(ddof=1)) if len(s) > 1 else np.nan
    fig = go.Figure()
    fig.add_trace(go.Histogram(x=s, nbinsx=26, name="Datos observados", opacity=0.78, marker=dict(line=dict(width=1, color="rgba(15,23,42,.25)"))))
    if len(s) > 1 and sigma > 0:
        x_min = min(float(s.min()), lie) - sigma
        x_max = max(float(s.max()), lse) + sigma
        x = np.linspace(x_min, x_max, 350)
        y = stats.norm.pdf(x, media, sigma)
        escala = len(s) * (x_max - x_min) / 26
        fig.add_trace(go.Scatter(x=x, y=y * escala, mode="lines", name="Aproximación normal", line=dict(width=4)))
    vline(fig, lie, "LIE")
    vline(fig, lse, "LSE")
    vline(fig, media, "Media", "solid")
    if vn is not None:
        vline(fig, vn, "VN", "dot")
    fig.update_xaxes(title="Variable de calidad")
    fig.update_yaxes(title="Frecuencia")
    return estilo_fig(fig, "Capacidad del proceso", "Histograma, normal aproximada, LIE, LSE y VN")


def grafico_escenario_mejora(go, media_actual, sigma_actual, media_nueva, sigma_objetivo, lie, lse, vn=None):
    if sigma_actual <= 0 or sigma_objetivo <= 0:
        return None
    minimo = min(lie, media_actual - 4 * sigma_actual, media_nueva - 4 * sigma_objetivo)
    maximo = max(lse, media_actual + 4 * sigma_actual, media_nueva + 4 * sigma_objetivo)
    x = np.linspace(minimo, maximo, 500)
    y_actual = stats.norm.pdf(x, media_actual, sigma_actual)
    y_nueva = stats.norm.pdf(x, media_nueva, sigma_objetivo)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=y_actual, mode="lines", name="Proceso actual", line=dict(width=4)))
    fig.add_trace(go.Scatter(x=x, y=y_nueva, mode="lines", name="Escenario mejorado", line=dict(width=4, dash="dash")))
    vline(fig, lie, "LIE")
    vline(fig, lse, "LSE")
    vline(fig, media_actual, "Media actual", "solid")
    vline(fig, media_nueva, "Nueva media", "dot")
    if vn is not None:
        vline(fig, vn, "VN", "dashdot")
    fig.update_xaxes(title="Variable de calidad")
    fig.update_yaxes(title="Densidad normal")
    return estilo_fig(fig, "Escenario de mejora", "Compara el proceso actual contra sigma objetivo y nueva media")


def grafico_xbar_r(go, df, col_valor, col_subgrupo):
    calc = calcular_xbar_r(df, col_valor, col_subgrupo)
    if calc is None:
        return None, None, None
    resumen, lim = calc
    fig_x = go.Figure()
    fig_x.add_trace(go.Scatter(x=resumen.index.astype(str), y=resumen["mean"], mode="lines+markers", name="Media del subgrupo", line=dict(width=3), marker=dict(size=8)))
    hline(fig_x, lim["LC Xbarra"], "LC X̄", "solid")
    hline(fig_x, lim["LCS Xbarra"], "LCS X̄")
    hline(fig_x, lim["LCI Xbarra"], "LCI X̄")
    fig_x.update_xaxes(title="Subgrupo")
    fig_x.update_yaxes(title="Media")
    fig_x = estilo_fig(fig_x, "Carta X-barra", "Control de la media del proceso")
    fig_r = go.Figure()
    fig_r.add_trace(go.Scatter(x=resumen.index.astype(str), y=resumen["R"], mode="lines+markers", name="Rango", line=dict(width=3), marker=dict(size=8)))
    hline(fig_r, lim["LC R"], "LC R", "solid")
    hline(fig_r, lim["LCS R"], "LCS R")
    hline(fig_r, lim["LCI R"], "LCI R")
    fig_r.update_xaxes(title="Subgrupo")
    fig_r.update_yaxes(title="Rango")
    fig_r = estilo_fig(fig_r, "Carta R", "Control de la variabilidad dentro del subgrupo")
    return fig_x, fig_r, lim


def grafico_xbar_s(go, df, col_valor, col_subgrupo):
    calc = calcular_xbar_s(df, col_valor, col_subgrupo)
    if calc is None:
        return None, None, None
    resumen, lim = calc
    fig_x = go.Figure()
    fig_x.add_trace(go.Scatter(x=resumen.index.astype(str), y=resumen["mean"], mode="lines+markers", name="Media del subgrupo", line=dict(width=3), marker=dict(size=8)))
    hline(fig_x, lim["LC Xbarra"], "LC X̄", "solid")
    hline(fig_x, lim["LCS Xbarra"], "LCS X̄")
    hline(fig_x, lim["LCI Xbarra"], "LCI X̄")
    fig_x.update_xaxes(title="Subgrupo")
    fig_x.update_yaxes(title="Media")
    fig_x = estilo_fig(fig_x, "Carta X-barra", "Control de la media del proceso con carta S")
    fig_s = go.Figure()
    fig_s.add_trace(go.Scatter(x=resumen.index.astype(str), y=resumen["std"], mode="lines+markers", name="Desviación estándar", line=dict(width=3), marker=dict(size=8)))
    hline(fig_s, lim["LC S"], "LC S", "solid")
    hline(fig_s, lim["LCS S"], "LCS S")
    hline(fig_s, lim["LCI S"], "LCI S")
    fig_s.update_xaxes(title="Subgrupo")
    fig_s.update_yaxes(title="S")
    fig_s = estilo_fig(fig_s, "Carta S", "Control de la variabilidad dentro del subgrupo")
    return fig_x, fig_s, lim


def grafico_i_mr(go, serie):
    calc = calcular_i_mr(serie)
    if calc is None:
        return None, None, None
    s, mr, lim = calc
    fig_i = go.Figure()
    fig_i.add_trace(go.Scatter(x=s.index + 1, y=s, mode="lines+markers", name="Individual", line=dict(width=3), marker=dict(size=8)))
    hline(fig_i, lim["LC I"], "LC I", "solid")
    hline(fig_i, lim["LCS I"], "LCS I")
    hline(fig_i, lim["LCI I"], "LCI I")
    fig_i.update_xaxes(title="Observación")
    fig_i.update_yaxes(title="Valor individual")
    fig_i = estilo_fig(fig_i, "Carta I", "Valores individuales")
    fig_mr = go.Figure()
    fig_mr.add_trace(go.Scatter(x=mr.index + 1, y=mr, mode="lines+markers", name="Rango móvil", line=dict(width=3), marker=dict(size=8)))
    hline(fig_mr, lim["LC MR"], "LC MR", "solid")
    hline(fig_mr, lim["LCS MR"], "LCS MR")
    hline(fig_mr, lim["LCI MR"], "LCI MR")
    fig_mr.update_xaxes(title="Observación")
    fig_mr.update_yaxes(title="Rango móvil")
    fig_mr = estilo_fig(fig_mr, "Carta MR", "Rango móvil entre observaciones consecutivas")
    return fig_i, fig_mr, lim


def grafico_lineas_atributos(go, x, y, lc=None, lcs=None, lci=None, titulo="Carta de control", subtitulo=""):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=y, mode="lines+markers", name="Valor", line=dict(width=3), marker=dict(size=8)))
    if lc is not None:
        hline(fig, lc, "LC", "solid")
    if lcs is not None:
        if np.isscalar(lcs):
            hline(fig, lcs, "LCS")
        else:
            fig.add_trace(go.Scatter(x=x, y=lcs, mode="lines", name="LCS", line=dict(width=3, dash="dash")))
    if lci is not None:
        if np.isscalar(lci):
            hline(fig, lci, "LCI")
        else:
            fig.add_trace(go.Scatter(x=x, y=lci, mode="lines", name="LCI", line=dict(width=3, dash="dash")))
    fig.update_xaxes(title="Muestra")
    fig.update_yaxes(title="Valor de control")
    return estilo_fig(fig, titulo, subtitulo)


def grafico_oc(go, datos, aql=None, ltpd=None):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=datos["Fracción defectuosa"], y=datos["Probabilidad de aceptación"], mode="lines", name="Curva OC", line=dict(width=4)))
    if aql is not None:
        vline(fig, aql, "AQL / NAC")
    if ltpd is not None:
        vline(fig, ltpd, "LTPD / NRC")
    fig.update_xaxes(title="Fracción defectuosa")
    fig.update_yaxes(title="Probabilidad de aceptación")
    return estilo_fig(fig, "Curva característica de operación", "Riesgo del productor y consumidor")


def crear_hoja_vacia(filas: int = 25, columnas_medicion: int = 5) -> pd.DataFrame:
    """Crea una hoja editable en formato ancho: filas=subgrupos y columnas=mediciones."""
    filas = max(1, int(filas))
    columnas_medicion = max(1, int(columnas_medicion))
    datos = {"#MUESTRAS": list(range(1, filas + 1))}
    for i in range(1, columnas_medicion + 1):
        datos[str(i)] = [np.nan] * filas
    return pd.DataFrame(datos)


def ajustar_dimension_hoja(df: pd.DataFrame, filas: int, columnas_medicion: int) -> pd.DataFrame:
    """Ajusta la hoja editable al número solicitado de filas y columnas sin perder datos existentes cuando sea posible."""
    filas = max(1, int(filas))
    columnas_medicion = max(1, int(columnas_medicion))
    columnas_objetivo = ["#MUESTRAS"] + [str(i) for i in range(1, columnas_medicion + 1)]
    salida = pd.DataFrame(index=range(filas), columns=columnas_objetivo)
    salida["#MUESTRAS"] = list(range(1, filas + 1))
    if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
        df_temp = df.copy().reset_index(drop=True)
        for col in columnas_objetivo:
            if col in df_temp.columns:
                limite = min(filas, len(df_temp))
                salida.loc[:limite - 1, col] = df_temp.loc[:limite - 1, col].values
        salida["#MUESTRAS"] = salida["#MUESTRAS"].fillna(pd.Series(range(1, filas + 1)))
    return salida


def dataframe_hoja_ejemplo() -> pd.DataFrame:
    """Datos de prueba en formato ancho para validar cartas X-barra/R."""
    return pd.DataFrame({
        "#MUESTRAS": list(range(1, 26)),
        "1": [308.1,302.6,300.3,307.9,308.9,302.2,302.4,305.7,306.2,312.1,302.7,313.2,302.6,299.1,301.3,307.1,315.5,301.8,304.8,305.1,309.6,306.7,303.1,298.2,310.4],
        "2": [304.8,305.3,301.3,308.8,307.0,310.3,302.7,316.0,306.9,308.9,303.6,304.5,297.8,299.8,304.9,312.6,299.5,309.3,294.5,302.1,298.5,300.9,310.6,307.2,304.5],
        "3": [294.8,299.1,300.8,304.6,304.9,301.8,304.0,307.1,304.1,313.5,299.1,299.4,306.0,301.5,304.2,305.4,310.4,300.3,314.0,307.6,306.7,304.9,303.5,308.9,303.5],
        "4": [296.6,308.9,305.4,306.7,306.3,301.4,298.2,310.7,312.4,306.8,303.9,311.2,309.1,307.5,303.3,305.8,306.8,307.1,304.0,303.5,307.2,297.1,297.6,311.9,293.4],
        "5": [309.1,309.0,305.1,309.5,304.7,300.6,303.8,309.4,300.8,307.1,302.3,310.0,304.7,304.8,302.8,319.1,315.0,306.2,312.1,307.5,301.0,306.6,309.1,303.5,299.6],
    })


def normalizar_decimales_coma(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte valores con coma decimal a números cuando sea posible."""
    salida = df.copy()
    for col in salida.columns:
        if salida[col].dtype == object:
            serie = salida[col].astype(str).str.strip().str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
            convertida = pd.to_numeric(serie, errors="coerce")
            if convertida.notna().sum() > 0:
                salida[col] = convertida
    return salida


def convertir_hoja_ancha_a_largo(df: pd.DataFrame, columna_subgrupo: str) -> pd.DataFrame:
    """Convierte tabla tipo Excel: filas=subgrupos, columnas=mediciones, a formato largo."""
    temp = normalizar_decimales_coma(df).dropna(how="all").copy()
    if columna_subgrupo not in temp.columns:
        raise ValueError("La columna de subgrupo no existe.")
    columnas_medicion = [c for c in temp.columns if c != columna_subgrupo]
    temp = temp.melt(id_vars=columna_subgrupo, value_vars=columnas_medicion, var_name="medicion_id", value_name="medicion")
    temp = temp.rename(columns={columna_subgrupo: "subgrupo"})
    temp["medicion"] = pd.to_numeric(temp["medicion"], errors="coerce")
    temp = temp.dropna(subset=["medicion"])
    return normalizar_nombres_columnas(temp)



# ============================================================
# PANTALLAS
# ============================================================

def pantalla_inicio(st):
    tarjetas(st, {"Estado": "Listo", "Enfoque": "CEP / SPC", "Secciones": 8, "Creador": "Jerson López"}, "Panel principal")
    caja_estado(st, "info", "Pulso de Calidad SPC integra gestión de datos, gráficos de control, supuestos, capacidad, diseño ARL/ATS, muestreo de aceptación y reportes. Está organizado según el alcance técnico del trabajo final.")

    st.markdown("### Mapa de módulos del software")
    modulos = plan_modulos_df()
    st.dataframe(modulos, use_container_width=True, hide_index=True)
    boton_exportar_excel(st, {"Mapa de secciones": modulos}, "mapa_secciones_pulso_calidad.xlsx", "Descargar mapa de secciones")

    st.markdown("### Ruta recomendada de análisis")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        caja_estado(st, "info", "1. Carga, limpia y caracteriza los datos.")
        caja_estado(st, "info", "2. Define variable crítica y subgrupos racionales.")
    with c2:
        caja_estado(st, "info", "3. Valida supuestos y estabilidad preliminar.")
        caja_estado(st, "info", "4. Construye cartas en Fase I y monitorea en Fase II.")
    with c3:
        caja_estado(st, "info", "5. Evalúa capacidad, PNC y aptitud del proceso.")
        caja_estado(st, "info", "6. Diseña ARL/ATS y plan de muestreo.")
    with c4:
        caja_estado(st, "info", "7. Genera conclusiones con juicio de ingeniería.")
        caja_estado(st, "info", "8. Exporta resultados para informe y anexos.")

    st.markdown("### Diccionario de conceptos")
    dic = pd.DataFrame({
        "Sigla": ["VN", "LIE", "LSE", "Cp", "CPU", "CPL", "Cpk", "PNC", "ARL0", "ARL1", "ATS"],
        "Nombre completo": ["Valor nominal u objetivo", "Límite inferior de especificación", "Límite superior de especificación", "Capacidad potencial", "Capacidad hacia LSE", "Capacidad hacia LIE", "Capacidad real", "Producto no conforme", "Muestras promedio hasta falsa alarma", "Muestras promedio hasta detectar cambio", "Tiempo promedio hasta señal"],
    })
    st.dataframe(dic, use_container_width=True, hide_index=True)
    boton_exportar_excel(st, {"Diccionario": dic}, "diccionario_spc.xlsx", "Descargar diccionario")


def pantalla_cargar_datos(st):
    st.markdown("### Gestión de datos")
    caja_estado(st, "info", "Carga datos desde Excel o CSV, pega una tabla en formato CSV, limpia el archivo, identifica variables y define subgrupos racionales para el análisis.")

    tab_archivo, tab_resumen = st.tabs(["Cargar datos", "Resumen, limpieza y subgrupos"])

    with tab_archivo:
        st.markdown("#### Carga de archivo")
        archivo = st.file_uploader("Sube un archivo CSV o Excel", type=["csv", "xlsx"], key="archivo_principal")
        if archivo is not None:
            try:
                df = normalizar_decimales_coma(cargar_datos_desde_archivo(archivo))
                st.session_state["df"] = df
                st.session_state["df_original"] = df.copy()
                actualizar_parametros_desde_df(st, df)
                st.success("Datos cargados correctamente.")
                st.dataframe(df.head(80), use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"Error al cargar archivo: {e}")

        st.markdown("#### Carga manual en formato CSV")
        caja_estado(st, "info", "Pega una tabla con encabezados. Se aceptan separadores por coma o punto y coma. Para decimales con coma, la app realiza conversión automática cuando es posible.")
        texto = st.text_area(
            "Pega datos con encabezados",
            height=170,
            placeholder="subgrupo;medicion_1;medicion_2;medicion_3\n1;308,1;304,8;294,8\n2;302,6;305,3;299,1",
            key="texto_csv_principal",
        )
        sep = st.radio("Separador del texto pegado", ["Automático", "Coma (,)", "Punto y coma (;)", "Tabulación"], horizontal=True, key="sep_texto_manual")
        if st.button("Cargar datos pegados", key="btn_cargar_csv_pegado"):
            try:
                if sep == "Coma (,)":
                    separador = ","
                elif sep == "Punto y coma (;)":
                    separador = ";"
                elif sep == "Tabulación":
                    separador = "\t"
                else:
                    separador = None

                if sep == "Coma (,)":
                    df = pd.read_csv(StringIO(texto), sep=",")
                elif sep == "Punto y coma (;)":
                    df = pd.read_csv(StringIO(texto), sep=";")
                elif sep == "Tabulación":
                    df = pd.read_csv(StringIO(texto), sep="\t")
                else:
                    try:
                        df = pd.read_csv(StringIO(texto), sep=None, engine="python")
                    except Exception:
                        df = pd.read_csv(StringIO(texto))

                df = normalizar_decimales_coma(normalizar_nombres_columnas(df))
                st.session_state["df"] = df
                st.session_state["df_original"] = df.copy()
                actualizar_parametros_desde_df(st, df)
                st.success("Datos pegados cargados correctamente.")
                st.dataframe(df.head(80), use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"No se pudieron cargar los datos: {e}")

    with tab_resumen:
        df = st.session_state.get("df")
        if not dataframe_valido(df):
            st.warning("Primero carga datos desde archivo o desde texto CSV.")
            return
        panel_parametros_activos(st)
        tarjetas(st, {"Filas": len(df), "Columnas": len(df.columns), "Numéricas": len(columnas_numericas(df)), "Faltantes": int(df.isna().sum().sum())}, "Resumen del archivo")
        st.dataframe(df.head(80), use_container_width=True, hide_index=True)
        with st.expander("Convertir formato ancho a formato largo"):
            caja_estado(st, "info", "Usa esta opción cuando tus datos vienen con una fila por subgrupo y varias columnas de medición. El resultado queda como: subgrupo, medicion_id y medicion.")
            col_sub_conv = st.selectbox("Columna que identifica el subgrupo", df.columns.tolist(), key="conv_col_subgrupo")
            if st.button("Convertir", key="btn_convertir_ancho_largo"):
                try:
                    st.session_state["df"] = convertir_hoja_ancha_a_largo(df, col_sub_conv)
                    actualizar_parametros_desde_df(st, st.session_state["df"])
                    st.success("Datos convertidos a formato largo.")
                    st.dataframe(st.session_state["df"], use_container_width=True, hide_index=True)
                except Exception as e:
                    st.error(f"No se pudo convertir: {e}")
        st.markdown("### Resumen estadístico")
        resumen_general = df.describe(include="all").T.reset_index().rename(columns={"index": "Variable"})
        st.dataframe(resumen_general, use_container_width=True, hide_index=True)
        st.markdown("### Perfil de variables y limpieza")
        perfil = perfil_datos(df)
        st.dataframe(perfil, use_container_width=True, hide_index=True)
        c_limp1, c_limp2, c_limp3 = st.columns(3)
        with c_limp1:
            if st.button("Limpiar filas vacías", key="btn_limpiar_filas"):
                st.session_state["df"] = df.dropna(how="all").copy()
                actualizar_parametros_desde_df(st, st.session_state["df"])
                st.success("Filas vacías eliminadas.")
                st.rerun()
        with c_limp2:
            if st.button("Eliminar duplicados exactos", key="btn_duplicados"):
                st.session_state["df"] = df.drop_duplicates().copy()
                actualizar_parametros_desde_df(st, st.session_state["df"])
                st.success("Duplicados exactos eliminados.")
                st.rerun()
        with c_limp3:
            boton_exportar_excel(st, {"Datos": df, "Perfil": perfil, "Resumen": resumen_general}, "gestion_datos_resumen.xlsx")
        st.markdown("### Definición de subgrupos racionales")
        caja_estado(st, "info", "Un subgrupo racional reúne mediciones tomadas bajo condiciones similares. Úsalo para cartas X-barra/R o X-barra/S. Si no hay subgrupos, usa I-MR.")
        cols_sub = ["Ninguno"] + df.columns.tolist()
        sub_sugerido = "subgrupo" if "subgrupo" in df.columns else cols_sub[0]
        sub_def = selectbox_persistente(st, "Columna sugerida de subgrupo", cols_sub, "datos_subgrupo_racional", sub_sugerido)
        if sub_def != "Ninguno":
            conteo_sub = df.groupby(sub_def).size().reset_index(name="Tamaño del subgrupo")
            st.dataframe(conteo_sub.head(50), use_container_width=True, hide_index=True)
            boton_exportar_excel(st, {"Subgrupos": conteo_sub}, "subgrupos_racionales.xlsx", "Descargar subgrupos")

def pantalla_supuestos(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    col = selectbox_persistente(st, "Variable de medición", nums, "sup_variable", nums[0])
    s = convertir_a_numerica(obtener_columna(df, col))
    tarjetas(st, resumen_descriptivo(s), "Resumen de la variable")
    tab1, tab2, tab3, tab4 = st.tabs(["Normalidad", "Independencia", "Homocedasticidad", "Atípicos"])

    with tab1:
        norm = evaluar_normalidad(s)
        estado, texto = diagnostico_normalidad(norm)
        caja_estado(st, "ok" if estado == "Cumple" else "alerta" if estado == "No cumple" else "info", texto)
        st.dataframe(norm, use_container_width=True, hide_index=True)
        a, b = st.columns(2)
        with a:
            st.plotly_chart(grafico_qq(go, s), use_container_width=True, key="supuestos_qq")
        with b:
            st.plotly_chart(grafico_box(go, s), use_container_width=True, key="box_supuestos_normalidad")

    with tab2:
        ind = evaluar_independencia(s)
        tarjetas(st, ind, "Independencia temporal")
        st.plotly_chart(grafico_corrida(go, s), use_container_width=True, key="supuestos_corrida")

    with tab3:
        opciones_grupo = ["Ninguna"] + df.columns.tolist()
        grupo = selectbox_persistente(st, "Columna de grupo", opciones_grupo, "sup_grupo", "Ninguna")
        if grupo != "Ninguna":
            st.dataframe(evaluar_homocedasticidad(df, col, grupo), use_container_width=True, hide_index=True)
        else:
            caja_estado(st, "info", "Selecciona una columna de grupo para evaluar homocedasticidad.")

    with tab4:
        li, ls, atip = detectar_atipicos_iqr(s)
        tarjetas(st, {"Límite IQR inferior": li, "Límite IQR superior": ls, "Atípicos": len(atip)}, "Detección IQR")
        st.plotly_chart(grafico_box(go, s), use_container_width=True, key="box_supuestos_atipicos")
        if len(atip) > 0:
            st.dataframe(atip.reset_index(), use_container_width=True, hide_index=True)

    try:
        sup_export = {
            "Resumen": pd.DataFrame([resumen_descriptivo(s)]).T.reset_index().rename(columns={"index": "Indicador", 0: "Valor"}),
            "Normalidad": evaluar_normalidad(s),
            "Independencia": pd.DataFrame([evaluar_independencia(s)]),
            "Atípicos": atip.reset_index() if 'atip' in locals() else pd.DataFrame(),
        }
        boton_exportar_excel(st, sup_export, "validacion_supuestos.xlsx", "Descargar validación de supuestos")
    except Exception:
        pass


def _tabla_patrones_control(valores, lc, lcs, lci):
    try:
        return evaluar_patrones_shewhart(valores, lc, lcs, lci)
    except Exception:
        return reglas_shewhart(valores, lc, lcs, lci)


def _graficar_fase_ii_linea(go, x, y, lc, lcs, lci, titulo, subtitulo, ytitle="Valor"):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=y, mode="lines+markers", name="Datos Fase II", line=dict(width=3), marker=dict(size=8)))
    hline(fig, lc, "LC Fase I", "solid")
    hline(fig, lcs, "LCS Fase I", "dash")
    hline(fig, lci, "LCI Fase I", "dash")
    fig.update_xaxes(title="Muestra / subgrupo")
    fig.update_yaxes(title=ytitle)
    return estilo_fig(fig, titulo, subtitulo)


def _mostrar_resumen_fase(st, lim, titulo):
    datos = {k: v for k, v in lim.items() if isinstance(v, (int, float, np.integer, np.floating, str))}
    tarjetas(st, datos, titulo)


def _leer_datos_fase_ii(st):
    df_base = st.session_state.get("df")
    st.markdown("##### Entrada de datos nuevos para Fase II")
    caja_estado(st, "info", "En Fase II se cargan datos nuevos del proceso. La app usa los límites guardados de Fase I y no recalcula los límites base.")
    archivo = st.file_uploader("Cargar nuevos datos Fase II desde Excel o CSV", type=["xlsx", "csv"], key="fase2_archivo_nuevo")
    texto = st.text_area(
        "O pega datos nuevos en formato CSV",
        height=110,
        placeholder="subgrupo,medicion\n26,304.2\n26,305.1\n27,306.0",
        key="fase2_csv_manual",
    )
    usar_base = st.checkbox("Usar los mismos datos cargados como práctica de monitoreo", value=False, key="fase2_usar_base")

    df2 = None
    if archivo is not None:
        try:
            df2 = cargar_datos_desde_archivo(archivo)
            st.success("Datos nuevos de Fase II cargados desde archivo.")
        except Exception as e:
            st.error(f"No se pudieron leer los datos nuevos del archivo: {e}")
            return None
    elif texto.strip():
        try:
            df2 = normalizar_nombres_columnas(pd.read_csv(StringIO(texto)))
            st.success("Datos nuevos de Fase II cargados desde texto.")
        except Exception as e:
            st.error(f"No se pudieron leer los datos nuevos: {e}")
            return None
    elif usar_base:
        df2 = df_base.copy() if dataframe_valido(df_base) else None

    if dataframe_valido(df2):
        st.caption("Vista previa de datos nuevos Fase II")
        st.dataframe(df2.head(25), use_container_width=True, hide_index=True)
        boton_exportar_excel(st, {"Datos Fase II": df2, "Perfil Fase II": perfil_datos(df2)}, "datos_fase_ii.xlsx", "Descargar datos Fase II")
        return df2
    return None


def pantalla_control(st, go):
    st.markdown("### Gráficos de control por Fase I y Fase II")
    caja_estado(
        st,
        "info",
        "Este módulo está separado explícitamente en Fase I y Fase II. Primero se estiman límites con datos históricos. Luego se usan esos límites fijos para monitorear datos nuevos."
    )

    modo = radio_persistente(
        st,
        "Selecciona la fase de trabajo",
        ["Fase I: estimar límites", "Fase II: monitorear con límites de Fase I", "Vista completa"],
        "ctrl_modo_fase",
        "Vista completa",
        horizontal=True,
    )

    c_f1, c_f2 = st.columns(2)
    with c_f1:
        caja_estado(st, "ok", "Fase I: usa datos históricos, calcula LC, LCS y LCI, revisa estabilidad y permite guardar los límites base.")
    with c_f2:
        caja_estado(st, "info", "Fase II: usa límites guardados de Fase I, evalúa nuevas muestras y genera alertas sin recalcular los límites base.")

    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos en el módulo Cargar datos. La selección de Fase I/Fase II ya está activa aquí.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    filtro = st.session_state.get("filtro_control_tipo", "todos")
    if filtro == "variables":
        opciones_carta = ["I-MR", "X-barra y R", "X-barra y S"]
        caja_estado(st, "info", "Selección justificada: usa I-MR para mediciones individuales; X-barra/R para subgrupos de 2 a 10; X-barra/S para subgrupos mayores o cuando interesa controlar la desviación estándar.")
    elif filtro == "atributos":
        opciones_carta = ["p", "np", "c", "u"]
        caja_estado(st, "info", "Selección justificada: p para proporción no conforme con n variable; np para número no conforme con n constante; c para defectos con oportunidad constante; u para defectos por unidad con oportunidad variable.")
    else:
        opciones_carta = ["I-MR", "X-barra y R", "X-barra y S", "p", "np", "c", "u"]

    valor_defecto_carta = st.session_state.get("ctrl_tipo", opciones_carta[0])
    if valor_defecto_carta not in opciones_carta:
        valor_defecto_carta = opciones_carta[0]
    tipo = radio_persistente(
        st,
        "Tipo de carta",
        opciones_carta,
        "ctrl_tipo",
        valor_defecto_carta,
        horizontal=True,
    )

    def guardar_limites(nombre, lim):
        st.session_state[nombre] = lim
        caja_estado(st, "ok", "Límites de Fase I guardados. Ya puedes usarlos en Fase II.")

    if tipo == "I-MR":
        col = selectbox_persistente(st, "Variable continua", nums, "ctrl_imr_variable", nums[0])
        calc = calcular_i_mr(obtener_columna(df, col))
        if calc is None:
            st.error("No se pudo calcular I-MR. Revisa que haya variación y al menos 3 datos.")
            return
        s_val, mr, lim = calc
        if modo in ["Fase I: estimar límites", "Vista completa"]:
            fig_i, fig_mr, _ = grafico_i_mr(go, obtener_columna(df, col))
            _mostrar_resumen_fase(st, lim, "Fase I: límites estimados I-MR")
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(fig_i, use_container_width=True, key="fase1_imr_i")
            with c2:
                st.plotly_chart(fig_mr, use_container_width=True, key="fase1_imr_mr")
            st.dataframe(_tabla_patrones_control(s_val, lim["LC I"], lim["LCS I"], lim["LCI I"]), use_container_width=True, hide_index=True)
            if st.button("Guardar límites I-MR para Fase II"):
                guardar_limites("lim_fase1_imr", lim)
        if modo in ["Fase II: monitorear con límites de Fase I", "Vista completa"]:
            st.markdown("#### Fase II: monitoreo I-MR")
            lim2 = st.session_state.get("lim_fase1_imr", lim)
            df2 = _leer_datos_fase_ii(st)
            if dataframe_valido(df2) and col in df2.columns:
                s2 = convertir_a_numerica(obtener_columna(df2, col)).reset_index(drop=True)
                if len(s2) >= 1:
                    fig2 = _graficar_fase_ii_linea(go, s2.index + 1, s2, lim2["LC I"], lim2["LCS I"], lim2["LCI I"], "Fase II - Carta I", "Límites fijos estimados en Fase I", col)
                    st.plotly_chart(fig2, use_container_width=True, key="fase2_imr_i")
                    fuera = int(((s2 > lim2["LCS I"]) | (s2 < lim2["LCI I"])).sum())
                    tarjetas(st, {"Puntos monitoreados": len(s2), "Puntos fuera": fuera, "Estado": "Alerta" if fuera else "Sin señal"}, "Resultado Fase II")
                    st.dataframe(_tabla_patrones_control(s2, lim2["LC I"], lim2["LCS I"], lim2["LCI I"]), use_container_width=True, hide_index=True)
            else:
                caja_estado(st, "alerta", "Para Fase II pega datos nuevos con la misma columna de medición o activa el uso de datos cargados.")

    elif tipo in ["X-barra y R", "X-barra y S"]:
        col = selectbox_persistente(st, "Variable continua", nums, "ctrl_xbar_variable", nums[0])
        sub = selectbox_persistente(st, "Subgrupo racional", df.columns.tolist(), "ctrl_xbar_subgrupo", df.columns.tolist()[0])
        usa_s = tipo == "X-barra y S"
        calc = calcular_xbar_s(df, col, sub) if usa_s else calcular_xbar_r(df, col, sub)
        if calc is None:
            st.error("No se pudo calcular la carta. Revisa que cada subgrupo tenga datos suficientes y tamaño compatible.")
            return
        resumen, lim = calc
        nombre_estado = "lim_fase1_xbars" if usa_s else "lim_fase1_xbarr"
        if modo in ["Fase I: estimar límites", "Vista completa"]:
            fig_x, fig_v, _ = grafico_xbar_s(go, df, col, sub) if usa_s else grafico_xbar_r(go, df, col, sub)
            _mostrar_resumen_fase(st, lim, f"Fase I: límites estimados {tipo}")
            c1, c2 = st.columns(2)
            with c1:
                st.plotly_chart(fig_x, use_container_width=True, key=f"fase1_{tipo}_x")
            with c2:
                st.plotly_chart(fig_v, use_container_width=True, key=f"fase1_{tipo}_v")
            st.markdown("#### Subgrupos usados en Fase I")
            st.dataframe(resumen.reset_index(), use_container_width=True, hide_index=True)
            st.markdown("#### Reglas sobre la carta de medias")
            st.dataframe(_tabla_patrones_control(resumen["mean"], lim["LC Xbarra"], lim["LCS Xbarra"], lim["LCI Xbarra"]), use_container_width=True, hide_index=True)
            if st.button(f"Guardar límites {tipo} para Fase II"):
                guardar_limites(nombre_estado, lim)
        if modo in ["Fase II: monitorear con límites de Fase I", "Vista completa"]:
            st.markdown(f"#### Fase II: monitoreo {tipo}")
            lim2 = st.session_state.get(nombre_estado, lim)
            df2 = _leer_datos_fase_ii(st)
            if dataframe_valido(df2) and col in df2.columns and sub in df2.columns:
                temp = pd.DataFrame({"valor": pd.to_numeric(obtener_columna(df2, col), errors="coerce"), "subgrupo": obtener_columna(df2, sub)}).dropna()
                if not temp.empty:
                    res2 = temp.groupby("subgrupo")["valor"].agg(["mean", "max", "min", "std", "count"])
                    res2 = res2[res2["count"] >= 2]
                    if not res2.empty:
                        res2["R"] = res2["max"] - res2["min"]
                        fig2 = _graficar_fase_ii_linea(go, res2.index.astype(str), res2["mean"], lim2["LC Xbarra"], lim2["LCS Xbarra"], lim2["LCI Xbarra"], "Fase II - Carta X-barra", "Límites fijos estimados en Fase I", "Media")
                        st.plotly_chart(fig2, use_container_width=True, key=f"fase2_{tipo}_x")
                        fuera = int(((res2["mean"] > lim2["LCS Xbarra"]) | (res2["mean"] < lim2["LCI Xbarra"])).sum())
                        tarjetas(st, {"Subgrupos monitoreados": len(res2), "Puntos fuera X-barra": fuera, "Estado": "Alerta" if fuera else "Sin señal"}, "Resultado Fase II")
                        st.dataframe(res2.reset_index(), use_container_width=True, hide_index=True)
            else:
                caja_estado(st, "alerta", "Para Fase II pega datos nuevos con las mismas columnas de variable y subgrupo.")

    elif tipo in ["p", "np"]:
        dcol = selectbox_persistente(st, "Unidades defectuosas / no conformes", nums, f"ctrl_{tipo}_defectuosos", nums[0])
        ncol = selectbox_persistente(st, "Unidades inspeccionadas", nums, f"ctrl_{tipo}_inspeccionados", nums[0])
        calc = calcular_p(df, dcol, ncol) if tipo == "p" else calcular_np(df, dcol, ncol)
        if calc is None:
            st.error(f"No se pudo calcular carta {tipo}.")
            return
        temp, lim = calc
        if modo in ["Fase I: estimar límites", "Vista completa"]:
            _mostrar_resumen_fase(st, lim, f"Fase I: carta {tipo}")
            if tipo == "p":
                st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["p"], lim["p barra"], temp["LCS"], temp["LCI"], "Carta p - Fase I", "Proporción no conforme"), use_container_width=True, key="fase1_p")
            else:
                st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["def"], lim["LC np"], lim["LCS np"], lim["LCI np"], "Carta np - Fase I", "Número de unidades no conformes"), use_container_width=True, key="fase1_np")
            if st.button(f"Guardar límites carta {tipo} para Fase II"):
                guardar_limites(f"lim_fase1_{tipo}", lim)
        if modo in ["Fase II: monitorear con límites de Fase I", "Vista completa"]:
            st.markdown(f"#### Fase II: monitoreo carta {tipo}")
            lim2 = st.session_state.get(f"lim_fase1_{tipo}", lim)
            df2 = _leer_datos_fase_ii(st)
            if dataframe_valido(df2) and dcol in df2.columns and ncol in df2.columns:
                calc2 = calcular_p(df2, dcol, ncol) if tipo == "p" else calcular_np(df2, dcol, ncol)
                if calc2:
                    temp2, _ = calc2
                    if tipo == "p":
                        pbar = lim2["p barra"]
                        temp2["LCS_FaseI"] = (pbar + 3 * np.sqrt(pbar * (1 - pbar) / temp2["n"])).clip(upper=1)
                        temp2["LCI_FaseI"] = (pbar - 3 * np.sqrt(pbar * (1 - pbar) / temp2["n"])).clip(lower=0)
                        st.plotly_chart(grafico_lineas_atributos(go, temp2.index + 1, temp2["p"], pbar, temp2["LCS_FaseI"], temp2["LCI_FaseI"], "Carta p - Fase II", "Límites fijos de Fase I"), use_container_width=True, key="fase2_p")
                        fuera = int(((temp2["p"] > temp2["LCS_FaseI"]) | (temp2["p"] < temp2["LCI_FaseI"])).sum())
                    else:
                        st.plotly_chart(grafico_lineas_atributos(go, temp2.index + 1, temp2["def"], lim2["LC np"], lim2["LCS np"], lim2["LCI np"], "Carta np - Fase II", "Límites fijos de Fase I"), use_container_width=True, key="fase2_np")
                        fuera = int(((temp2["def"] > lim2["LCS np"]) | (temp2["def"] < lim2["LCI np"])).sum())
                    tarjetas(st, {"Muestras monitoreadas": len(temp2), "Puntos fuera": fuera, "Estado": "Alerta" if fuera else "Sin señal"}, "Resultado Fase II")
                    st.dataframe(temp2.reset_index(), use_container_width=True, hide_index=True)

    elif tipo in ["c", "u"]:
        dcol = selectbox_persistente(st, "Defectos", nums, f"ctrl_{tipo}_defectos", nums[0])
        ncol = None
        if tipo == "u":
            ncol = selectbox_persistente(st, "Unidades u oportunidades inspeccionadas", nums, "ctrl_u_unidades", nums[0])
        calc = calcular_c(df, dcol) if tipo == "c" else calcular_u(df, dcol, ncol)
        if calc is None:
            st.error(f"No se pudo calcular carta {tipo}.")
            return
        temp, lim = calc
        if modo in ["Fase I: estimar límites", "Vista completa"]:
            _mostrar_resumen_fase(st, lim, f"Fase I: carta {tipo}")
            if tipo == "c":
                st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["c"], lim["LC c"], lim["LCS c"], lim["LCI c"], "Carta c - Fase I", "Número de defectos"), use_container_width=True, key="fase1_c")
            else:
                st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["u"], lim["u barra"], temp["LCS"], temp["LCI"], "Carta u - Fase I", "Defectos por unidad"), use_container_width=True, key="fase1_u")
            if st.button(f"Guardar límites carta {tipo} para Fase II"):
                guardar_limites(f"lim_fase1_{tipo}", lim)
        if modo in ["Fase II: monitorear con límites de Fase I", "Vista completa"]:
            st.markdown(f"#### Fase II: monitoreo carta {tipo}")
            lim2 = st.session_state.get(f"lim_fase1_{tipo}", lim)
            df2 = _leer_datos_fase_ii(st)
            if dataframe_valido(df2) and dcol in df2.columns and (tipo == "c" or ncol in df2.columns):
                calc2 = calcular_c(df2, dcol) if tipo == "c" else calcular_u(df2, dcol, ncol)
                if calc2:
                    temp2, _ = calc2
                    if tipo == "c":
                        st.plotly_chart(grafico_lineas_atributos(go, temp2.index + 1, temp2["c"], lim2["LC c"], lim2["LCS c"], lim2["LCI c"], "Carta c - Fase II", "Límites fijos de Fase I"), use_container_width=True, key="fase2_c")
                        fuera = int(((temp2["c"] > lim2["LCS c"]) | (temp2["c"] < lim2["LCI c"])).sum())
                    else:
                        ubar = lim2["u barra"]
                        temp2["LCS_FaseI"] = ubar + 3 * np.sqrt(ubar / temp2["n"])
                        temp2["LCI_FaseI"] = (ubar - 3 * np.sqrt(ubar / temp2["n"])).clip(lower=0)
                        st.plotly_chart(grafico_lineas_atributos(go, temp2.index + 1, temp2["u"], ubar, temp2["LCS_FaseI"], temp2["LCI_FaseI"], "Carta u - Fase II", "Límites fijos de Fase I"), use_container_width=True, key="fase2_u")
                        fuera = int(((temp2["u"] > temp2["LCS_FaseI"]) | (temp2["u"] < temp2["LCI_FaseI"])).sum())
                    tarjetas(st, {"Muestras monitoreadas": len(temp2), "Puntos fuera": fuera, "Estado": "Alerta" if fuera else "Sin señal"}, "Resultado Fase II")
                    st.dataframe(temp2.reset_index(), use_container_width=True, hide_index=True)

def pantalla_capacidad(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return

    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Análisis de capacidad del proceso")
    caja_estado(
        st,
        "info",
        "Los valores quedan guardados aunque cambies de módulo. Este apartado no incluye boxplot para hacerlo más liviano."
    )

    variable = selectbox_persistente(st, "Variable crítica de calidad", nums, "cap_variable", nums[0])
    serie = convertir_a_numerica(obtener_columna(df, variable))
    if len(serie) < 2:
        st.error("La variable seleccionada necesita al menos 2 datos numéricos.")
        return

    media_base = float(serie.mean())
    sigma_base = float(serie.std(ddof=1)) if len(serie) > 1 else 1.0
    if pd.isna(sigma_base) or sigma_base <= 0:
        sigma_base = 1.0

    p_activos = parametros_proceso(st)
    if p_activos.get("variable") != variable:
        actualizar_parametros_desde_df(st, df, variable)
        p_activos = parametros_proceso(st)
    media_base = float(p_activos.get("media", media_base))
    sigma_base = float(p_activos.get("sigma", sigma_base))
    vn_base = float(p_activos.get("vn", media_base))
    lie_base = float(p_activos.get("lie", serie.min()))
    lse_base = float(p_activos.get("lse", serie.max()))
    tolerancia_base = float(p_activos.get("tolerancia", max(1.0, sigma_base)))

    guardar_estado_si_no_existe(st, "cap_vn", vn_base)
    guardar_estado_si_no_existe(st, "cap_usar_tolerancia", True)
    guardar_estado_si_no_existe(st, "cap_tolerancia", tolerancia_base)
    guardar_estado_si_no_existe(st, "cap_lie", lie_base)
    guardar_estado_si_no_existe(st, "cap_lse", lse_base)
    guardar_estado_si_no_existe(st, "cap_metodo_sigma", "Sigma muestral")
    guardar_estado_si_no_existe(st, "cap_sigma_historica", sigma_base)
    guardar_estado_si_no_existe(st, "cap_rbarra", 10.100)
    guardar_estado_si_no_existe(st, "cap_d2", 2.326)
    guardar_estado_si_no_existe(st, "cap_usar_media_conocida", False)
    guardar_estado_si_no_existe(st, "cap_media_conocida", media_base)
    guardar_estado_si_no_existe(st, "cap_cp_objetivo", 1.33)
    guardar_estado_si_no_existe(st, "cap_pnc_lse_permitido", 0.05)
    guardar_estado_si_no_existe(st, "cap_media_detectar", media_base)
    guardar_estado_si_no_existe(st, "cap_nsub", 5)
    guardar_estado_si_no_existe(st, "cap_intervalo", 15.0)
    guardar_estado_si_no_existe(st, "cap_potencia_objetivo", 0.90)

    col_config, col_resumen = st.columns([1.05, 0.95])

    with col_config:
        vn = numero_persistente(st, "VN | Valor nominal u objetivo", "cap_vn", vn_base)
        usar_tolerancia = checkbox_persistente(
            st,
            "Calcular LIE y LSE desde VN ± tolerancia",
            "cap_usar_tolerancia",
            True,
        )

        if usar_tolerancia:
            tolerancia = numero_persistente(
                st,
                "Tolerancia",
                "cap_tolerancia",
                tolerancia_base,
                min_value=0.000001,
            )
            lie = vn - tolerancia
            lse = vn + tolerancia
            st.write(f"LIE = {fmt(lie)} | LSE = {fmt(lse)}")
        else:
            lie = numero_persistente(
                st,
                "LIE | Límite inferior de especificación",
                "cap_lie",
                lie_base,
            )
            lse = numero_persistente(
                st,
                "LSE | Límite superior de especificación",
                "cap_lse",
                lse_base,
            )

        metodo_sigma = radio_persistente(
            st,
            "Método de sigma",
            ["Sigma muestral", "Sigma histórica", "R̄/d2"],
            "cap_metodo_sigma",
            "Sigma muestral",
            horizontal=True,
        )

        sigma_historica = None
        if metodo_sigma == "Sigma histórica":
            sigma_historica = numero_persistente(
                st,
                "Sigma histórica",
                "cap_sigma_historica",
                sigma_base,
                min_value=0.000001,
                format="%.8f",
            )
        elif metodo_sigma == "R̄/d2":
            rbarra = numero_persistente(
                st,
                "R̄ | Rango promedio",
                "cap_rbarra",
                10.100,
                min_value=0.000001,
            )
            d2 = numero_persistente(
                st,
                "d2",
                "cap_d2",
                2.326,
                min_value=0.000001,
            )
            sigma_historica = rbarra / d2
            st.write(f"Sigma estimada = {fmt(sigma_historica)}")

        usar_media = checkbox_persistente(
            st,
            "Usar media conocida",
            "cap_usar_media_conocida",
            False,
        )
        media_conocida = None
        if usar_media:
            media_conocida = numero_persistente(
                st,
                "Media conocida",
                "cap_media_conocida",
                media_base,
                format="%.8f",
            )

    with col_resumen:
        tarjetas(st, resumen_descriptivo(serie), "Resumen de la variable")
        caja_estado(st, "info", "Para revisar boxplot y atípicos entra a Supuestos del proceso → Atípicos.")

    if lie >= lse:
        st.error("LIE debe ser menor que LSE.")
        return

    resultado = calcular_capacidad(serie, lie, lse, vn, sigma_historica, media_conocida)
    if resultado is None:
        st.error("No se pudo calcular capacidad. Revisa la variabilidad, LIE y LSE.")
        return

    guardar_parametros_proceso(
        st,
        variable=variable,
        media=resultado["Media usada"],
        sigma=resultado["Sigma usada"],
        vn=resultado["VN"],
        lie=resultado["LIE"],
        lse=resultado["LSE"],
        tolerancia=max(abs(resultado["LSE"] - resultado["VN"]), abs(resultado["VN"] - resultado["LIE"])),
        cp=resultado["Cp"],
        cpk=resultado["Cpk"],
        cpu=resultado["CPU"],
        cpl=resultado["CPL"],
        pnc_lie=resultado["% PNC estimado LIE"],
        pnc_lse=resultado["% PNC estimado LSE"],
        pnc_total=resultado["% PNC estimado total"],
        riesgo=resultado["Riesgo principal"],
        estado=resultado["Estado"],
    )
    sincronizar_parametros_widgets(st)
    panel_parametros_activos(st)

    tarjetas(st, {
        "Cp": resultado["Cp"],
        "CPL": resultado["CPL"],
        "CPU": resultado["CPU"],
        "Cpk": resultado["Cpk"],
        "PNC estimado total %": resultado["% PNC estimado total"],
        "Estado": resultado["Estado"],
        "Centrado": resultado["Centrado"],
        "Riesgo": resultado["Riesgo principal"],
    }, "Indicadores de capacidad")

    t1, t2, t3, t4, t5 = st.tabs([
        "Gráfica",
        "Tabla completa",
        "Producto no conforme",
        "Escenario de mejora",
        "Potencia X-barra",
    ])

    with t1:
        st.plotly_chart(
            grafico_hist_capacidad(go, serie, lie, lse, vn),
            use_container_width=True,
            key="capacidad_hist",
        )

    with t2:
        tabla(st, resultado, "Capacidad completa")

    with t3:
        tarjetas(st, {
            "Fuera por LIE observado": resultado["Fuera por LIE observado"],
            "Fuera por LSE observado": resultado["Fuera por LSE observado"],
            "% PNC LIE estimado": resultado["% PNC estimado LIE"],
            "% PNC LSE estimado": resultado["% PNC estimado LSE"],
            "% PNC total estimado": resultado["% PNC estimado total"],
            "PPM estimado": resultado["PPM estimado"],
        }, "Producto no conforme")

    with t4:
        st.markdown("#### Simulación de mejora del proceso")
        cp_objetivo = numero_persistente(
            st,
            "Cp objetivo",
            "cap_cp_objetivo",
            1.33,
            min_value=0.1,
            step=0.01,
        )
        pnc_lse_permitido = numero_persistente(
            st,
            "PNC máximo permitido hacia LSE",
            "cap_pnc_lse_permitido",
            0.05,
            min_value=0.000001,
            max_value=0.5,
            step=0.001,
            format="%.8f",
        )

        sigma_objetivo = calcular_sigma_objetivo(lie, lse, cp_objetivo)
        nueva_media = media_maxima_para_pnc_lse(lse, sigma_objetivo, pnc_lse_permitido)
        capacidad_nueva = calcular_capacidad(
            pd.Series([nueva_media - sigma_objetivo, nueva_media, nueva_media + sigma_objetivo]),
            lie,
            lse,
            vn,
            sigma_objetivo,
            nueva_media,
        )

        tarjetas(st, {
            "Sigma actual": resultado["Sigma usada"],
            "Sigma objetivo": sigma_objetivo,
            "Media actual": resultado["Media usada"],
            "Nueva media sugerida": nueva_media,
            "PNC permitido hacia LSE %": pnc_lse_permitido * 100,
            "Cpk nuevo estimado": capacidad_nueva["Cpk"] if capacidad_nueva else np.nan,
            "PNC total nuevo %": capacidad_nueva["% PNC estimado total"] if capacidad_nueva else np.nan,
        }, "Resultados del escenario")

        fig_mejora = grafico_escenario_mejora(
            go,
            resultado["Media usada"],
            resultado["Sigma usada"],
            nueva_media,
            sigma_objetivo,
            lie,
            lse,
            vn,
        )
        if fig_mejora is not None:
            st.plotly_chart(
                fig_mejora,
                use_container_width=True,
                key="capacidad_escenario_mejora",
            )
        caja_estado(
            st,
            "info",
            "La curva continua representa el proceso actual. La curva punteada representa el proceso con sigma objetivo y nueva media sugerida."
        )

    with t5:
        st.markdown("#### Potencia para carta X-barra")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            cambio = numero_persistente(
                st,
                "Media que se desea detectar",
                "cap_media_detectar",
                float(vn),
            )
        with c2:
            nsub = numero_persistente(
                st,
                "Tamaño de subgrupo",
                "cap_nsub",
                5,
                min_value=1,
                step=1,
            )
        with c3:
            intervalo = numero_persistente(
                st,
                "Intervalo entre muestras en minutos",
                "cap_intervalo",
                15.0,
                min_value=1.0,
            )
        with c4:
            zcrit = numero_persistente(
                st,
                "Z crítico",
                "cap_zcrit",
                3.0,
                min_value=0.01,
                step=0.1,
            )

        pot = potencia_xbarra(resultado["Media usada"], cambio, resultado["Sigma usada"], int(nsub), float(zcrit))
        if pot:
            pot["ATS1 minutos"] = pot["ARL1"] * intervalo if np.isfinite(pot["ARL1"]) else np.inf
            tabla(st, pot, "Potencia X-barra detallada")
            pasos = [
                f"1. Media actual = {fmt(pot['Media actual'])}.",
                f"2. Media a detectar = {fmt(pot['Media a detectar'])}.",
                f"3. Sigma usada = {fmt(pot['Sigma usada'])}.",
                f"4. Diferencia absoluta = |media a detectar - media actual| = {fmt(pot['Diferencia absoluta'])}.",
                f"5. d = diferencia absoluta / sigma usada = {fmt(pot['d'])}.",
                f"6. d√n = d × raíz del tamaño de subgrupo = {fmt(pot['d√n'])}.",
                f"7. Z crítico - d√n = {fmt(pot['Z crítico - d√n'])}.",
                f"8. β = P(no detectar el cambio) = {fmt(pot['β'])}, equivalente a {fmt(pot['% β'])}%.",
                f"9. Potencia = 1 - β = {fmt(pot['Potencia = 1 - β'])}, equivalente a {fmt(pot['% Potencia'])}%.",
                f"10. ARL1 = 1 / potencia = {fmt(pot['ARL1'])}. ATS1 = ARL1 × intervalo = {fmt(pot['ATS1 minutos'])} minutos.",
            ]
            st.markdown("#### Explicación paso a paso")
            for paso in pasos:
                caja_estado(st, "info", paso)

        objetivo_pot = numero_persistente(
            st,
            "Potencia objetivo",
            "cap_potencia_objetivo",
            0.90,
            min_value=0.01,
            max_value=0.99,
        )
        n_req = n_para_potencia(resultado["Media usada"], cambio, resultado["Sigma usada"], objetivo_pot, float(zcrit))
        if n_req:
            caja_estado(st, "ok", f"Tamaño de subgrupo sugerido: n = {n_req}")

    st.markdown("### Diagnóstico")
    for mensaje in diagnostico_capacidad(resultado):
        tipo = "ok" if "cumple" in mensaje.lower() or "no se observaron" in mensaje.lower() else "alerta"
        caja_estado(st, tipo, mensaje)


def pantalla_no_conformes(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    modo = radio_persistente(
        st,
        "Tipo",
        ["Por especificación", "Pareto por categoría"],
        "nc_modo",
        "Por especificación",
        horizontal=True,
    )

    if modo == "Por especificación":
        p_activos = parametros_proceso(st)
        variable_defecto = p_activos.get("variable", nums[0]) if p_activos.get("variable", nums[0]) in nums else nums[0]
        col = selectbox_persistente(st, "Variable de medición", nums, "nc_variable", variable_defecto)
        s = convertir_a_numerica(obtener_columna(df, col))
        if p_activos.get("variable") != col:
            actualizar_parametros_desde_df(st, df, col)
            p_activos = parametros_proceso(st)
        guardar_estado_si_no_existe(st, "nc_vn", float(p_activos.get("vn", s.mean())))
        guardar_estado_si_no_existe(st, "nc_lie", float(p_activos.get("lie", s.min())))
        guardar_estado_si_no_existe(st, "nc_lse", float(p_activos.get("lse", s.max())))

        c1, c2, c3 = st.columns(3)
        with c1:
            vn = numero_persistente(st, "VN", "nc_vn", float(s.mean()))
        with c2:
            lie = numero_persistente(st, "LIE", "nc_lie", float(s.min()))
        with c3:
            lse = numero_persistente(st, "LSE", "nc_lse", float(s.max()))

        valores = pd.to_numeric(obtener_columna(df, col), errors="coerce")
        n_validos = len(valores.dropna())
        bajo = df[valores < lie]
        alto = df[valores > lse]
        total = df[(valores < lie) | (valores > lse)]
        pct_lie = len(bajo) / n_validos * 100 if n_validos else 0
        pct_lse = len(alto) / n_validos * 100 if n_validos else 0
        pct_total = len(total) / n_validos * 100 if n_validos else 0
        tarjetas(st, {
            "Evaluados": n_validos,
            "Conformes": n_validos - len(total),
            "No conformes": len(total),
            "Por LIE": len(bajo),
            "% por LIE": pct_lie,
            "Por LSE": len(alto),
            "% por LSE": pct_lse,
            "% no conforme total": pct_total,
        }, "Resumen no conforme")
        detalle_nc = pd.DataFrame([
            {"Límite": "LIE", "Condición": f"{col} < {fmt(lie)}", "No conformes": len(bajo), "% sobre evaluados": pct_lie},
            {"Límite": "LSE", "Condición": f"{col} > {fmt(lse)}", "No conformes": len(alto), "% sobre evaluados": pct_lse},
            {"Límite": "Total", "Condición": "Fuera de LIE o LSE", "No conformes": len(total), "% sobre evaluados": pct_total},
        ])
        st.markdown("### Porcentaje de no conformes por límite")
        st.dataframe(detalle_nc, use_container_width=True, hide_index=True)
        guardar_parametros_proceso(st, variable=col, vn=vn, lie=lie, lse=lse, pnc_observado=pct_total, pnc_observado_lie=pct_lie, pnc_observado_lse=pct_lse)
        sincronizar_parametros_widgets(st)
        st.plotly_chart(grafico_hist_capacidad(go, s, lie, lse, vn), use_container_width=True, key="no_conformes_hist")
        t1, t2, t3 = st.tabs(["Todos", "Bajo LIE", "Sobre LSE"])
        with t1:
            st.dataframe(total, use_container_width=True, hide_index=True)
        with t2:
            st.dataframe(bajo, use_container_width=True, hide_index=True)
        with t3:
            st.dataframe(alto, use_container_width=True, hide_index=True)
    else:
        cat = selectbox_persistente(st, "Columna de categoría", df.columns.tolist(), "nc_categoria", df.columns.tolist()[0])
        conteo = df[cat].astype(str).value_counts().reset_index()
        conteo.columns = ["Categoría", "Frecuencia"]
        conteo["%"] = conteo["Frecuencia"] / conteo["Frecuencia"].sum() * 100
        conteo["% acumulado"] = conteo["%"].cumsum()
        fig = go.Figure()
        fig.add_trace(go.Bar(x=conteo["Categoría"], y=conteo["Frecuencia"], name="Frecuencia"))
        fig.add_trace(go.Scatter(x=conteo["Categoría"], y=conteo["% acumulado"], mode="lines+markers", name="% acumulado", yaxis="y2", line=dict(width=4)))
        fig.update_layout(yaxis2=dict(overlaying="y", side="right", range=[0, 100], title="% acumulado"))
        st.plotly_chart(estilo_fig(fig, "Pareto de no conformidades", "Frecuencia y porcentaje acumulado"), use_container_width=True, key="pareto_no_conformes")
        st.dataframe(conteo, use_container_width=True, hide_index=True)


def pantalla_muestreo(st, go):
    st.markdown("### Muestreo de aceptación")
    guardar_estado_si_no_existe(st, "muestreo_n", 50)
    guardar_estado_si_no_existe(st, "muestreo_c", 2)
    guardar_estado_si_no_existe(st, "muestreo_aql", 0.025)
    guardar_estado_si_no_existe(st, "muestreo_ltpd", 0.05)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        n = numero_persistente(st, "n | Tamaño de muestra", "muestreo_n", 50, min_value=1, step=1)
    with c2:
        c = numero_persistente(st, "c | Número de aceptación", "muestreo_c", 2, min_value=0, step=1)
    with c3:
        aql = numero_persistente(st, "AQL / NAC", "muestreo_aql", 0.025, min_value=0.0001, max_value=1.0, step=0.001, format="%.4f")
    with c4:
        ltpd = numero_persistente(st, "LTPD / NRC", "muestreo_ltpd", 0.05, min_value=0.0001, max_value=1.0, step=0.001, format="%.4f")

    if c > n:
        st.error("c no puede ser mayor que n.")
        return
    datos = curva_oc(int(n), int(c))
    pa_aql = float(stats.binom.cdf(int(c), int(n), float(aql)))
    pa_ltpd = float(stats.binom.cdf(int(c), int(n), float(ltpd)))
    tarjetas(st, {"Pa(AQL)": pa_aql, "Alfa": 1 - pa_aql, "Pa(LTPD) / Beta": pa_ltpd, "Plan": f"n={int(n)}, c={int(c)}"}, "Plan de muestreo")
    st.plotly_chart(grafico_oc(go, datos, aql, ltpd), use_container_width=True, key="muestreo_oc")


def pantalla_diseno_graficos(st, go):
    st.markdown("### Diseño de gráficos ARL y ATS")
    caja_estado(st, "info", "Diseña una carta X-barra según la rapidez con la que debe detectar cambios en la media. Calcula potencia, beta, ARL0, ATS0, ARL1, ATS1 y tamaño de subgrupo recomendado.")

    p_activos = parametros_proceso(st)
    media_def = float(p_activos.get("media", 0.0))
    sigma_def = float(p_activos.get("sigma", 1.0))
    lse_def = float(p_activos.get("lse", media_def + sigma_def))
    n_def = int(p_activos.get("nsub", 5)) if p_activos.get("nsub", 5) else 5
    intervalo_def = float(p_activos.get("intervalo", 15.0))
    panel_parametros_activos(st)

    guardar_estado_si_no_existe(st, "dis_media_actual", media_def)
    guardar_estado_si_no_existe(st, "dis_media_cambio", lse_def)
    guardar_estado_si_no_existe(st, "dis_sigma", sigma_def)
    guardar_estado_si_no_existe(st, "dis_n", n_def)
    guardar_estado_si_no_existe(st, "dis_intervalo", intervalo_def)
    guardar_estado_si_no_existe(st, "dis_z", 3.0)
    guardar_estado_si_no_existe(st, "dis_pot_obj", 0.90)

    c1, c2, c3 = st.columns(3)
    with c1:
        media_actual = numero_persistente(st, "Media actual del proceso", "dis_media_actual", media_def, format="%.8f")
        media_cambio = numero_persistente(st, "Media que se desea detectar", "dis_media_cambio", lse_def, format="%.8f")
    with c2:
        sigma = numero_persistente(st, "Sigma del proceso", "dis_sigma", sigma_def, min_value=0.000001, format="%.8f")
        n = numero_persistente(st, "Tamaño de subgrupo n", "dis_n", n_def, min_value=1, step=1)
    with c3:
        intervalo = numero_persistente(st, "Tiempo entre muestras en minutos", "dis_intervalo", intervalo_def, min_value=0.000001, format="%.4f")
        z = numero_persistente(st, "Z crítico", "dis_z", 3.0, min_value=1.0, max_value=6.0, step=0.1)

    res = calcular_desempeno_grafico_xbarra(float(media_actual), float(media_cambio), float(sigma), int(n), float(intervalo), float(z))
    if res is None:
        st.error("No se pudo calcular el desempeño. Revisa sigma, n e intervalo.")
        return

    guardar_parametros_proceso(st, media=float(media_actual), sigma=float(sigma), media_detectar=float(media_cambio), nsub=int(n), intervalo=float(intervalo), arl0=res["ARL0"], arl1=res["ARL1"], ats0=res["ATS0 minutos"], ats1=res["ATS1 minutos"], potencia=res["potencia = 1 - β"])

    tarjetas(st, {
        "Potencia %": res["% potencia"],
        "β %": res["% β"],
        "ARL0": res["ARL0"],
        "ATS0 min": res["ATS0 minutos"],
        "ARL1": res["ARL1"],
        "ATS1 min": res["ATS1 minutos"],
        "d": res["d"],
        "d√n": res["d√n"],
    }, "Desempeño del gráfico")

    tab1, tab2, tab3 = st.tabs(["Tabla completa", "Diseñar n", "Explicación"])
    with tab1:
        tabla(st, res, "ARL, ATS y potencia")
    with tab2:
        pot_obj = numero_persistente(st, "Potencia objetivo", "dis_pot_obj", 0.90, min_value=0.01, max_value=0.999, step=0.01)
        n_req = n_para_potencia_xbarra(float(media_actual), float(media_cambio), float(sigma), float(pot_obj), float(z))
        if n_req:
            res_req = calcular_desempeno_grafico_xbarra(float(media_actual), float(media_cambio), float(sigma), int(n_req), float(intervalo), float(z))
            tarjetas(st, {"n recomendado": n_req, "Potencia lograda %": res_req["% potencia"], "ARL1": res_req["ARL1"], "ATS1 min": res_req["ATS1 minutos"]}, "Diseño sugerido")
            caja_estado(st, "ok", f"Para detectar el cambio de {fmt(media_actual)} a {fmt(media_cambio)} con potencia cercana a {fmt(float(pot_obj)*100)}%, se recomienda n = {n_req}.")
        else:
            caja_estado(st, "alerta", "No se puede calcular n porque no hay cambio entre medias o los datos no son válidos.")
    with tab3:
        pasos = [
            f"1. d = |media a detectar - media actual| / sigma = {fmt(res['d'])}.",
            f"2. d√n = {fmt(res['d√n'])}.",
            f"3. β es la probabilidad de no detectar el cambio cuando ya ocurrió.",
            f"4. Potencia = 1 - β = {fmt(res['potencia = 1 - β'])}, equivalente a {fmt(res['% potencia'])}%.",
            f"5. ARL1 = 1 / potencia = {fmt(res['ARL1'])} muestras promedio para detectar el cambio.",
            f"6. ATS1 = ARL1 × intervalo = {fmt(res['ATS1 minutos'])} minutos promedio de detección.",
        ]
        for paso in pasos:
            caja_estado(st, "info", paso)


def generar_diagnostico_integral(df, variable, lie, lse, vn=None, sigma_hist=None):
    s = convertir_a_numerica(obtener_columna(df, variable))
    cap = calcular_capacidad(s, lie, lse, vn, sigma_hist if sigma_hist and sigma_hist > 0 else None, None) if lie < lse else None
    normalidad = evaluar_normalidad(s)
    estado_norm, texto_norm = diagnostico_normalidad(normalidad)
    ind = evaluar_independencia(s)
    filas = []
    filas.append({"Componente": "Normalidad", "Resultado": estado_norm, "Interpretación": texto_norm})
    filas.append({"Componente": "Independencia", "Resultado": ind.get("Resultado"), "Interpretación": ind.get("Interpretación")})
    if cap:
        filas.append({"Componente": "Capacidad", "Resultado": cap["Estado"], "Interpretación": f"Cp={fmt(cap['Cp'])}, Cpk={fmt(cap['Cpk'])}. Riesgo principal: {cap['Riesgo principal']}."})
        filas.append({"Componente": "Producto no conforme", "Resultado": f"{fmt(cap['% PNC estimado total'])}%", "Interpretación": f"LIE: {fmt(cap['% PNC estimado LIE'])}%, LSE: {fmt(cap['% PNC estimado LSE'])}%."})
        if cap["Cpk"] < 1:
            accion = "El proceso no debe liberarse como capaz sin acciones de mejora. Se recomienda reducir variabilidad y ajustar centrado."
        elif cap["Cpk"] < 1.33:
            accion = "El proceso es marginal. Debe mantenerse monitoreo frecuente y plan de reducción de variabilidad."
        else:
            accion = "El proceso tiene desempeño aceptable, sujeto a que se mantenga bajo control estadístico."
        filas.append({"Componente": "Juicio de ingeniería", "Resultado": "Acción técnica", "Interpretación": accion})
    return pd.DataFrame(filas), cap


def pantalla_conclusiones_mejora(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Conclusiones y plan de mejora")
    caja_estado(st, "info", "Este módulo convierte los resultados estadísticos en diagnóstico empresarial, conclusiones técnicas y acciones de mejora. Sirve para cualquier producto, proceso o variable de calidad.")

    p_activos = parametros_proceso(st)
    variable_defecto = p_activos.get("variable", nums[0]) if p_activos.get("variable", nums[0]) in nums else nums[0]
    variable = selectbox_persistente(st, "Variable crítica de calidad", nums, "conc_variable", variable_defecto)
    s = convertir_a_numerica(obtener_columna(df, variable))
    if p_activos.get("variable") != variable:
        actualizar_parametros_desde_df(st, df, variable)
        p_activos = parametros_proceso(st)
    panel_parametros_activos(st)
    guardar_estado_si_no_existe(st, "conc_vn", float(p_activos.get("vn", s.mean())))
    guardar_estado_si_no_existe(st, "conc_lie", float(p_activos.get("lie", s.min())))
    guardar_estado_si_no_existe(st, "conc_lse", float(p_activos.get("lse", s.max())))
    guardar_estado_si_no_existe(st, "conc_sigma", float(p_activos.get("sigma", 0.0)))
    guardar_estado_si_no_existe(st, "conc_producto", "producto")
    guardar_estado_si_no_existe(st, "conc_empresa", "empresa")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        empresa = st.text_input("Empresa", value=st.session_state.get("conc_empresa", "empresa"), key="conc_empresa")
        producto = st.text_input("Producto", value=st.session_state.get("conc_producto", "producto"), key="conc_producto")
    with c2:
        vn = numero_persistente(st, "VN u objetivo", "conc_vn", float(p_activos.get("vn", s.mean())), format="%.6f")
    with c3:
        lie = numero_persistente(st, "LIE", "conc_lie", float(p_activos.get("lie", s.min())), format="%.6f")
    with c4:
        lse = numero_persistente(st, "LSE", "conc_lse", float(p_activos.get("lse", s.max())), format="%.6f")
        sigma_hist = numero_persistente(st, "Sigma histórica opcional", "conc_sigma", float(p_activos.get("sigma", 0.0)), min_value=0.0, format="%.6f")

    diagnostico, cap = generar_diagnostico_integral(df, variable, float(lie), float(lse), float(vn), float(sigma_hist))
    st.dataframe(diagnostico, use_container_width=True, hide_index=True)

    if cap:
        tarjetas(st, {"Cp": cap["Cp"], "Cpk": cap["Cpk"], "PNC total %": cap["% PNC estimado total"], "Riesgo": cap["Riesgo principal"], "Estado": cap["Estado"], "Centrado": cap["Centrado"]}, "Resumen ejecutivo")
        acciones = []
        if cap["Riesgo principal"] == "LSE":
            acciones.append("Revisar ajuste de dosificación, calibración o parámetros operativos para evitar valores por encima del límite superior.")
        elif cap["Riesgo principal"] == "LIE":
            acciones.append("Revisar ajustes que estén generando valores por debajo del límite inferior.")
        if cap["Cp"] < 1:
            acciones.append("Reducir variabilidad mediante mantenimiento, estandarización operativa, control de materia prima y capacitación.")
        if cap["Cpk"] < cap["Cp"] * 0.85:
            acciones.append("Centrar la media del proceso hacia el valor objetivo o centro de especificación.")
        acciones.append("Mantener seguimiento con cartas de control y documentar causas especiales cuando aparezcan señales fuera de control.")
        st.markdown("#### Plan de acción priorizado")
        plan = pd.DataFrame({"Prioridad": range(1, len(acciones)+1), "Acción recomendada": acciones, "Impacto esperado": ["Alto" if i < 3 else "Medio" for i in range(1, len(acciones)+1)]})
        st.dataframe(plan, use_container_width=True, hide_index=True)
        conclusion = (
            f"Con base en el análisis realizado en Pulso de Calidad SPC, el proceso asociado al producto {producto} en la empresa {empresa} "
            f"presenta una media de {fmt(cap['Media usada'])}, una sigma usada de {fmt(cap['Sigma usada'])}, Cp = {fmt(cap['Cp'])} y Cpk = {fmt(cap['Cpk'])}. "
            f"El proceso se clasifica como {cap['Estado']} y el riesgo principal se concentra hacia {cap['Riesgo principal']}. "
            f"El porcentaje estimado de producto no conforme total es {fmt(cap['% PNC estimado total'])}%, por lo que se recomienda priorizar acciones de mejora enfocadas en "
            f"reducir variabilidad, centrar el proceso y mantener control estadístico continuo."
        )
        st.markdown("#### Conclusión lista para informe")
        st.text_area("Puedes copiar este texto", value=conclusion, height=150)
    else:
        caja_estado(st, "alerta", "No se pudo calcular capacidad. Revisa LIE, LSE y variabilidad de los datos.")


def pantalla_asistente(st):
    st.markdown("### Asistente de proyecto")
    tipo = selectbox_persistente(
        st,
        "Tipo de dato",
        ["Variable continua", "Defectuoso / no defectuoso", "Número de defectos"],
        "asis_tipo",
        "Variable continua",
    )
    sub = checkbox_persistente(st, "Tengo subgrupos racionales", "asis_subgrupos", False)
    n = numero_persistente(st, "Tamaño del subgrupo", "asis_n", 5, min_value=1, step=1)
    variable_n = checkbox_persistente(st, "El tamaño de muestra cambia", "asis_n_variable", False)
    specs = checkbox_persistente(st, "Tengo LIE y LSE", "asis_specs", False)

    recs = []
    if tipo == "Variable continua":
        if sub and 2 <= n <= 10:
            recs.append("Usa carta X-barra/R porque tienes subgrupos racionales de tamaño entre 2 y 10.")
        elif sub and n > 10:
            recs.append("Para subgrupos mayores a 10 se recomienda X-barra/S. Si solo tienes esta app, revisa I-MR o transforma el análisis por subgrupos.")
        else:
            recs.append("Usa carta I-MR porque estás trabajando con mediciones individuales o sin subgrupos racionales.")
        if specs:
            recs.append("Calcula capacidad: Cp, CPU, CPL, Cpk, Pp, Ppk, Cpm y producto no conforme.")
    elif tipo == "Defectuoso / no defectuoso":
        if variable_n:
            recs.append("Usa carta p porque cambia el tamaño de muestra inspeccionado.")
        else:
            recs.append("Usa carta np porque el tamaño de muestra es constante.")
    else:
        if variable_n:
            recs.append("Usa carta u porque cambia el número de unidades u oportunidades inspeccionadas.")
        else:
            recs.append("Usa carta c porque el área de oportunidad es constante.")
    recs.append("Valida normalidad para capacidad de variables. Valida independencia antes de interpretar cartas.")
    for i, r in enumerate(recs, 1):
        caja_estado(st, "ok", f"{i}. {r}")


def pantalla_monitoreo(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Monitoreo en tiempo real")
    st.markdown("<span class='live-badge'>● Modo reproducción</span>", unsafe_allow_html=True)

    col = selectbox_persistente(st, "Variable", nums, "mon_variable", nums[0])
    guardar_estado_si_no_existe(st, "mon_ventana", min(25, len(df)))
    guardar_estado_si_no_existe(st, "mon_play", False)
    guardar_estado_si_no_existe(st, "mon_velocidad", 1)
    guardar_estado_si_no_existe(st, "mon_usar_specs", False)

    c1, c2, c3 = st.columns(3)
    with c1:
        ventana = numero_persistente(st, "Ventana visible", "mon_ventana", min(25, len(df)), min_value=5, max_value=min(100, len(df)), step=1)
    with c2:
        play = checkbox_persistente(st, "Reproducir como video", "mon_play", False)
        velocidad = numero_persistente(st, "Velocidad", "mon_velocidad", 1, min_value=1, max_value=5, step=1)
    with c3:
        usar = checkbox_persistente(st, "Agregar LIE y LSE", "mon_usar_specs", False)
        lie = lse = None
        if usar:
            s_temp = convertir_a_numerica(obtener_columna(df, col))
            guardar_estado_si_no_existe(st, "mon_lie", float(s_temp.min()))
            guardar_estado_si_no_existe(st, "mon_lse", float(s_temp.max()))
            lie = numero_persistente(st, "LIE", "mon_lie", float(s_temp.min()))
            lse = numero_persistente(st, "LSE", "mon_lse", float(s_temp.max()))

    s = convertir_a_numerica(obtener_columna(df, col)).reset_index(drop=True)
    if "monitor_pos" not in st.session_state:
        st.session_state["monitor_pos"] = min(int(ventana), len(s))
    if play:
        if st.session_state["monitor_pos"] >= len(s):
            st.session_state["monitor_pos"] = min(int(ventana), len(s))
        else:
            st.session_state["monitor_pos"] += 1
        time.sleep(float(velocidad))
        st.rerun()
    else:
        st.session_state["monitor_pos"] = st.slider("Línea de tiempo", min(int(ventana), len(s)), len(s), min(st.session_state["monitor_pos"], len(s)), key="mon_linea_tiempo")

    fin = st.session_state["monitor_pos"]
    ini = max(0, fin - int(ventana))
    sv = s.iloc[ini:fin]
    res_global = resumen_descriptivo(s)
    res_v = resumen_descriptivo(sv)
    ind = evaluar_independencia(s)
    st.progress(fin / len(s), text=f"Punto {fin} de {len(s)}")
    tarjetas(st, {"Último": float(sv.iloc[-1]), "Media global": res_global["Media"], "Media ventana": res_v["Media"], "Desv. ventana": res_v["Desv. estándar"], "Durbin-Watson": ind["Durbin-Watson"], "Autocorr. lag 1": ind["Autocorr. lag 1"], "Estado": ind["Resultado"]}, "Lectura actual")
    fig = grafico_corrida(go, sv, "Monitoreo dinámico")
    if usar and lie is not None and lse is not None and lie < lse:
        hline(fig, lie, "LIE", "dot")
        hline(fig, lse, "LSE", "dot")
    st.plotly_chart(fig, use_container_width=True, key="monitoreo_dinamico")
    st.dataframe(pd.DataFrame({"Orden": range(ini + 1, fin + 1), col: sv.values}), use_container_width=True, hide_index=True)


def escribir_titulo_informe(ws, fila, titulo):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    c = ws.cell(fila, 1)
    c.value = titulo
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = PatternFill("solid", fgColor="1D4ED8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 24
    return fila + 2


def escribir_tabla_informe(ws, fila, titulo, df):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    fila = escribir_titulo_informe(ws, fila, titulo)
    borde = Border(left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"), top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"))
    for col_idx, nombre in enumerate(df.columns, start=1):
        celda = ws.cell(fila, col_idx)
        celda.value = nombre
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F172A")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde
    for r_idx, row in enumerate(df.itertuples(index=False), start=fila + 1):
        for c_idx, valor in enumerate(row, start=1):
            celda = ws.cell(r_idx, c_idx)
            if isinstance(valor, (float, int, np.integer, np.floating)) and not pd.isna(valor):
                celda.value = valor_limpio(valor)
                celda.number_format = "General"
            else:
                celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(vertical="center", wrap_text=True)
    return fila + len(df) + 3


def ajustar_informe_excel(ws):
    from openpyxl.utils import get_column_letter
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    anchos = [26, 20, 20, 20, 20, 20, 20, 20]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def insertar_imagen_figura(ws, fig, celda, texto_fallo):
    try:
        from openpyxl.drawing.image import Image as XLImage
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp.close()
        fig.write_image(tmp.name, width=980, height=580, scale=1)
        img = XLImage(tmp.name)
        img.width = 680
        img.height = 400
        ws.add_image(img, celda)
        return True
    except Exception:
        ws[celda] = texto_fallo
        return False


def pantalla_reporte(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return

    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Informe visual en Excel")
    caja_estado(st, "info", "El informe se genera en una sola hoja. Usa los mismos gráficos de la app como imágenes. Para insertar imágenes instala: python -m pip install kaleido")
    panel_parametros_activos(st)

    c1, c2, c3 = st.columns(3)
    with c1:
        p_activos = parametros_proceso(st)
        variable_defecto = p_activos.get("variable", nums[0]) if p_activos.get("variable", nums[0]) in nums else nums[0]
        variable = selectbox_persistente(st, "Variable principal", nums, "rep_variable", variable_defecto)
        subgrupo = selectbox_persistente(st, "Subgrupo para carta X-barra/R", ["Ninguno"] + df.columns.tolist(), "rep_subgrupo", "Ninguno")
    with c2:
        guardar_estado_si_no_existe(st, "rep_vn", float(parametros_proceso(st).get("vn", convertir_a_numerica(obtener_columna(df, variable)).mean())))
        vn = numero_persistente(st, "VN | Valor nominal", "rep_vn", float(parametros_proceso(st).get("vn", convertir_a_numerica(obtener_columna(df, variable)).mean())))
        guardar_estado_si_no_existe(st, "rep_lie", float(parametros_proceso(st).get("lie", convertir_a_numerica(obtener_columna(df, variable)).min())))
        lie = numero_persistente(st, "LIE | Límite inferior", "rep_lie", float(parametros_proceso(st).get("lie", convertir_a_numerica(obtener_columna(df, variable)).min())))
    with c3:
        guardar_estado_si_no_existe(st, "rep_lse", float(parametros_proceso(st).get("lse", convertir_a_numerica(obtener_columna(df, variable)).max())))
        lse = numero_persistente(st, "LSE | Límite superior", "rep_lse", float(parametros_proceso(st).get("lse", convertir_a_numerica(obtener_columna(df, variable)).max())))
        guardar_estado_si_no_existe(st, "rep_sigma_hist", float(parametros_proceso(st).get("sigma", 0.0)))
        sigma_hist = numero_persistente(st, "Sigma histórica opcional", "rep_sigma_hist", float(parametros_proceso(st).get("sigma", 0.0)), min_value=0.0, format="%.6f")

    s = convertir_a_numerica(obtener_columna(df, variable))
    sigma_usada = sigma_hist if sigma_hist > 0 else None
    normalidad = evaluar_normalidad(s)
    estado_norm, texto_norm = diagnostico_normalidad(normalidad)
    resumen = pd.DataFrame([resumen_descriptivo(s)]).T.reset_index()
    resumen.columns = ["Indicador", "Valor"]
    cap = calcular_capacidad(s, lie, lse, vn, sigma_usada, None) if lie < lse else None

    if cap is not None:
        capacidad_df = pd.DataFrame([redondear_dict(cap)]).T.reset_index()
        capacidad_df.columns = ["Indicador", "Valor"]
        pnc_df = pd.DataFrame({
            "Indicador": ["% PNC estimado LIE", "% PNC estimado LSE", "% PNC estimado total", "PPM estimado", "PNC observado"],
            "Valor": [cap["% PNC estimado LIE"], cap["% PNC estimado LSE"], cap["% PNC estimado total"], cap["PPM estimado"], cap["Producto no conforme observado"]],
        })
    else:
        capacidad_df = pd.DataFrame({"Indicador": ["Capacidad"], "Valor": ["No calculada. Revisa LIE y LSE."]})
        pnc_df = pd.DataFrame({"Indicador": ["PNC"], "Valor": ["No calculado"]})

    config_df = pd.DataFrame({
        "Campo": ["Aplicación", "Variable", "Subgrupo", "VN", "LIE", "LSE", "Normalidad", "Fecha"],
        "Valor": ["Pulso de Calidad SPC", variable, subgrupo, vn, lie, lse, estado_norm, pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
    })

    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        pd.DataFrame({" ": []}).to_excel(writer, index=False, sheet_name="Informe")
        wb = writer.book
        ws = wb["Informe"]
        ws.delete_rows(1, 2)

        fila = 1
        fila = escribir_titulo_informe(ws, fila, "PULSO DE CALIDAD SPC - INFORME DE CONTROL ESTADÍSTICO")
        fila = escribir_tabla_informe(ws, fila, "Configuración del análisis", config_df)
        fila = escribir_tabla_informe(ws, fila, "Resumen estadístico", resumen)
        fila = escribir_tabla_informe(ws, fila, "Supuesto de normalidad", normalidad)
        ws.cell(fila, 1).value = "Diagnóstico de normalidad"
        ws.cell(fila, 2).value = texto_norm
        fila += 3
        fila = escribir_tabla_informe(ws, fila, "Análisis de capacidad", capacidad_df)
        fila = escribir_tabla_informe(ws, fila, "Producto no conforme", pnc_df)

        if subgrupo != "Ninguno":
            calc = calcular_xbar_r(df, variable, subgrupo)
            if calc is not None:
                resumen_x, lim_x = calc
                tabla_x = resumen_x.reset_index().rename(columns={"mean": "Xbarra"})
                fila = escribir_tabla_informe(ws, fila, "Datos de carta X-barra y R", tabla_x.head(35))
                fig_control, fig_r, _ = grafico_xbar_r(go, df, variable, subgrupo)
            else:
                fig_control = grafico_corrida(go, s, "Gráfico de corrida")
                fig_r = None
        else:
            figs = grafico_i_mr(go, s)
            fig_control = figs[0] if figs[0] is not None else grafico_corrida(go, s, "Gráfico de corrida")
            fig_r = figs[1] if figs[1] is not None else None

        fila = escribir_titulo_informe(ws, fila, "Gráficos generados desde la app")
        insertar_imagen_figura(ws, grafico_qq(go, s), f"A{fila}", "No se pudo insertar el gráfico de normalidad. Instala kaleido.")
        insertar_imagen_figura(ws, grafico_hist_capacidad(go, s, lie, lse, vn), f"A{fila + 23}", "No se pudo insertar el gráfico de capacidad. Instala kaleido.")
        insertar_imagen_figura(ws, fig_control, f"A{fila + 46}", "No se pudo insertar el gráfico de control. Instala kaleido.")
        if fig_r is not None:
            insertar_imagen_figura(ws, fig_r, f"A{fila + 69}", "No se pudo insertar la segunda carta de control. Instala kaleido.")
        ajustar_informe_excel(ws)

    tarjetas(st, {"Variable": variable, "Normalidad": estado_norm, "LIE": lie, "LSE": lse}, "Contenido del informe")
    st.download_button(
        "Descargar informe Excel",
        data=salida.getvalue(),
        file_name="informe_pulso_calidad_spc.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# PRUEBAS
# ============================================================

def ejecutar_pruebas():
    if FALTAN_BASE:
        print(mensaje_dependencias(FALTAN_BASE))
        return False
    s = pd.Series([10, 10.2, 9.8, 10.1, 9.9, 10.3, 10.0, 9.7])
    assert resumen_descriptivo(s)["n"] == 8
    assert evaluar_normalidad(s).shape[0] >= 1
    assert "Durbin-Watson" in evaluar_independencia(s)
    assert calcular_capacidad(s, 9, 11, 10) is not None
    df = pd.DataFrame({"subgrupo": [1, 1, 2, 2], "medicion": [10, 11, 9, 10]})
    assert calcular_xbar_r(df, "medicion", "subgrupo") is not None
    assert calcular_i_mr(s) is not None
    print("Todas las pruebas internas pasaron correctamente.")
    return True


# ============================================================
# APP
# ============================================================

def ejecutar_app():
    if FALTAN_BASE:
        print(mensaje_dependencias(FALTAN_BASE))
        return
    st, go = importar_librerias_app()
    if st is None or go is None:
        print(mensaje_dependencias(FALTAN_APP))
        return
    st.set_page_config(page_title="Pulso de Calidad SPC", page_icon="📊", layout="wide")
    aplicar_estilo_visual(st)
    if "df" not in st.session_state:
        st.session_state["df"] = None
    if "df_original" not in st.session_state:
        st.session_state["df_original"] = None
    st.sidebar.title("Pulso de Calidad SPC")
    st.sidebar.caption("Control Estadístico de Procesos")
    st.sidebar.caption("Desarrollador: Jerson Andrés López Wilches | Contacto: jerssonpriv@gmail.com")
    menu = st.sidebar.radio("Selecciona una sección", [
        "Inicio",
        "Gestión de datos",
        "Gráficos de control por variables",
        "Gráficos de control por atributos",
        "Validación de supuestos",
        "Capacidad del proceso",
        "Diseño de gráficos",
        "Muestreo de aceptación",
        "Reportes",
        "Conclusiones y plan de mejora",
        "Asistente de proyecto",
    ], key="menu_principal")
    encabezado(st, "Pulso de Calidad SPC", "Panel visual para analizar, controlar y mejorar procesos con datos.")
    
    if menu == "Inicio":
        pantalla_inicio(st)
    elif menu == "Gestión de datos":
        pantalla_cargar_datos(st)
    elif menu == "Gráficos de control por variables":
        st.info("Cartas por variables. Incluye I-MR, X-barra/R y X-barra/S, con Fase I para estimar límites y Fase II para monitorear con límites fijos.")
        st.session_state["filtro_control_tipo"] = "variables"
        pantalla_control(st, go)
    elif menu == "Gráficos de control por atributos":
        st.info("Cartas por atributos. Incluye p, np, c y u, con selección justificada según tipo de dato, tamaño muestral y naturaleza del defecto.")
        st.session_state["filtro_control_tipo"] = "atributos"
        pantalla_control(st, go)
    elif menu == "Validación de supuestos":
        pantalla_supuestos(st, go)
    elif menu == "Capacidad del proceso":
        pantalla_capacidad(st, go)
    elif menu == "Diseño de gráficos":
        pantalla_diseno_graficos(st, go)
    elif menu == "Muestreo de aceptación":
        pantalla_muestreo(st, go)
    elif menu == "Reportes":
        pantalla_reporte(st, go)
    elif menu == "Conclusiones y plan de mejora":
        pantalla_conclusiones_mejora(st, go)
    elif menu == "Asistente de proyecto":
        pantalla_asistente(st)


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        ejecutar_pruebas()
    else:
        ejecutar_app()
